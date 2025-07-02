from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, Count
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from .models import Transaccion, Categoria, Cuenta, Tag, ConfiguracionUsuario, Presupuesto, Meta, CorteMes, GrupoGastosCompartidos, GastoCompartido, PagoGastoCompartido, Notificacion
from .forms import (
    TransaccionForm, CategoriaForm, CuentaForm, TagForm, FiltroTransaccionesForm,
    PresupuestoForm, MetaForm, UsuarioCrearForm, UsuarioEditarForm, 
    ConfiguracionUsuarioForm, ConfiguracionSistemaForm, RegistroPublicoForm,
    GrupoGastosCompartidosForm, GastoCompartidoForm, PagoGastoCompartidoForm,
    FiltroGastosCompartidosForm
)
from datetime import date, datetime, timedelta
from decimal import Decimal
from django.utils import timezone
from calendar import monthrange
import json
from functools import wraps
from django.views.decorators.http import require_POST
from django.urls import reverse
import io
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from dateutil.relativedelta import relativedelta

# Decorador personalizado para verificar permisos de staff
def staff_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        if not (request.user.is_staff or request.user.is_superuser):
            messages.error(request, 'No tienes permisos para acceder a esta sección.')
            return render(request, 'configuracion/acceso_denegado.html', status=403)
        return view_func(request, *args, **kwargs)
    return _wrapped_view

def login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_active:
                auth_login(request, user)
                return redirect('dashboard')
            else:
                messages.error(request, 'Tu cuenta está pendiente de activación. Un administrador la revisará pronto.')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')
    
    return render(request, 'login/login.html')

def registro_publico(request):
    """Vista para registro público de usuarios (marcados como inactivos)"""
    if request.method == 'POST':
        form = RegistroPublicoForm(request.POST)
        if form.is_valid():
            user = form.save()
            
            # Crear configuración por defecto para el usuario
            ConfiguracionUsuario.objects.create(
                usuario=user,
                moneda_principal='ARS',
                zona_horaria='America/Argentina/Buenos_Aires',
                notificaciones_activas=True,
                recordatorios_pago=True
            )
            
            messages.success(request, 'Tu cuenta ha sido creada exitosamente. Un administrador la revisará y activará pronto.')
            return redirect('login')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = RegistroPublicoForm()
    
    return render(request, 'login/registro.html', {'form': form})

@login_required
def dashboard(request):
    # Obtener el mes actual
    hoy = date.today()
    inicio_mes = hoy.replace(day=1)
    
    # Estadísticas del mes
    ingresos_mes = Transaccion.objects.filter(
        usuario=request.user,
        tipo='ingreso',
        fecha__gte=inicio_mes
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
    
    gastos_mes = Transaccion.objects.filter(
        usuario=request.user,
        tipo='gasto',
        fecha__gte=inicio_mes
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
    
    balance_neto = ingresos_mes - gastos_mes
    
    # Cuentas activas
    cuentas_activas = Cuenta.objects.filter(usuario=request.user, activa=True)
    total_cuentas = cuentas_activas.count()
    
    # Calcular porcentaje de saldo para cada cuenta
    for cuenta in cuentas_activas:
        if cuenta.saldo_actual > 0:
            cuenta.porcentaje_saldo = min(100, (cuenta.saldo_actual / max(cuenta.saldo_actual, 1)) * 100)
        else:
            cuenta.porcentaje_saldo = abs(cuenta.saldo_actual) / max(abs(cuenta.saldo_actual), 1) * 100
    
    # Gastos por categoría del mes
    gastos_por_categoria = Transaccion.objects.filter(
        usuario=request.user,
        tipo='gasto',
        fecha__gte=inicio_mes
    ).values('categoria__nombre', 'categoria__color').annotate(
        total=Sum('monto')
    ).order_by('-total')
    
    # Calcular porcentajes para el gráfico
    total_gastos = sum(item['total'] for item in gastos_por_categoria)
    for item in gastos_por_categoria:
        if total_gastos > 0:
            item['porcentaje'] = (item['total'] / total_gastos) * 360  # Grados para el gráfico
        else:
            item['porcentaje'] = 0
    
    # Presupuestos activos
    presupuestos_activos = Presupuesto.objects.filter(
        usuario=request.user,
        estado='activo'
    ).order_by('-fecha_creacion')[:5]
    
    # Actualizar gastos de presupuestos
    for presupuesto in presupuestos_activos:
        presupuesto.actualizar_gasto()
    
    # Metas activas
    metas_activas = Meta.objects.filter(
        usuario=request.user,
        estado='activa'
    ).order_by('-fecha_creacion')[:5]
    
    # Transacciones recientes
    transacciones_recientes = Transaccion.objects.filter(
        usuario=request.user
    ).select_related('categoria', 'cuenta').order_by('-fecha', '-fecha_creacion')[:10]
    
    context = {
        'ingresos_mes': ingresos_mes,
        'gastos_mes': gastos_mes,
        'balance_neto': balance_neto,
        'total_cuentas': total_cuentas,
        'cuentas_activas': cuentas_activas,
        'gastos_por_categoria': gastos_por_categoria,
        'presupuestos_activos': presupuestos_activos,
        'metas_activas': metas_activas,
        'transacciones_recientes': transacciones_recientes,
    }
    
    return render(request, 'dashboard/dashboard.html', context)

def logout_view(request):
    auth_logout(request)
    return redirect('login')

# Vistas para Transacciones
@login_required(login_url='/finanzas/')
def transacciones_lista(request):
    form = FiltroTransaccionesForm(request.GET, user=request.user)
    transacciones = Transaccion.objects.filter(usuario=request.user).select_related('categoria', 'cuenta')
    
    # Aplicar filtros
    if form.is_valid():
        if form.cleaned_data.get('fecha_desde'):
            transacciones = transacciones.filter(fecha__gte=form.cleaned_data['fecha_desde'])
        if form.cleaned_data.get('fecha_hasta'):
            transacciones = transacciones.filter(fecha__lte=form.cleaned_data['fecha_hasta'])
        if form.cleaned_data.get('tipo'):
            transacciones = transacciones.filter(tipo=form.cleaned_data['tipo'])
        if form.cleaned_data.get('categoria'):
            transacciones = transacciones.filter(categoria=form.cleaned_data['categoria'])
        if form.cleaned_data.get('cuenta'):
            transacciones = transacciones.filter(cuenta=form.cleaned_data['cuenta'])
        if form.cleaned_data.get('monto_minimo'):
            transacciones = transacciones.filter(monto__gte=form.cleaned_data['monto_minimo'])
        if form.cleaned_data.get('monto_maximo'):
            transacciones = transacciones.filter(monto__lte=form.cleaned_data['monto_maximo'])
    
    # Paginación
    paginator = Paginator(transacciones, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_transacciones': transacciones.count(),
    }
    return render(request, 'transacciones/lista.html', context)

@login_required(login_url='/finanzas/')
def transaccion_crear(request):
    if request.method == 'POST':
        form = TransaccionForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            transaccion = form.save(commit=False)
            transaccion.usuario = request.user
            es_en_cuotas = form.cleaned_data.get('es_en_cuotas')
            numero_cuotas = form.cleaned_data.get('numero_cuotas')
            cuota_actual = form.cleaned_data.get('cuota_actual')
            fecha = form.cleaned_data.get('fecha')
            if es_en_cuotas and numero_cuotas and cuota_actual:
                # Crear todas las cuotas desde la actual hasta la última
                for i in range(cuota_actual, numero_cuotas + 1):
                    nueva_fecha = fecha + relativedelta(months=i - cuota_actual)
                    nueva_transaccion = Transaccion(
                        monto=transaccion.monto,
                        fecha=nueva_fecha,
                        tipo=transaccion.tipo,
                        descripcion=f"{transaccion.descripcion} (Cuota {i}/{numero_cuotas})",
                        categoria=transaccion.categoria,
                        cuenta=transaccion.cuenta,
                        usuario=request.user,
                        es_recurrente=transaccion.es_recurrente,
                        frecuencia_recurrencia=transaccion.frecuencia_recurrencia,
                        cuenta_destino=transaccion.cuenta_destino,
                        imagen_recibo=transaccion.imagen_recibo,
                        es_en_cuotas=True,
                        numero_cuotas=numero_cuotas,
                        cuota_actual=i,
                        fecha_fin_cuotas=transaccion.fecha_fin_cuotas
                    )
                    nueva_transaccion.save()
                    nueva_transaccion.tags.set(form.cleaned_data.get('tags'))
                messages.success(request, f'Transacción en cuotas creada exitosamente ({numero_cuotas} cuotas).')
                return redirect('transacciones_lista')
            else:
                transaccion.save()
                form.save_m2m()  # Guardar tags
                messages.success(request, 'Transacción creada exitosamente.')
                return redirect('transacciones_lista')
    else:
        form = TransaccionForm(user=request.user)
    
    context = {
        'form': form,
        'titulo': 'Nueva Transacción',
    }
    return render(request, 'transacciones/form.html', context)

@login_required(login_url='/finanzas/')
def transaccion_editar(request, pk):
    transaccion = get_object_or_404(Transaccion, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        form = TransaccionForm(request.POST, request.FILES, instance=transaccion, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Transacción actualizada exitosamente.')
            return redirect('transacciones_lista')
    else:
        form = TransaccionForm(instance=transaccion, user=request.user)
    
    context = {
        'form': form,
        'transaccion': transaccion,
        'titulo': 'Editar Transacción',
    }
    return render(request, 'transacciones/form.html', context)

@login_required(login_url='/finanzas/')
def transaccion_eliminar(request, pk):
    transaccion = get_object_or_404(Transaccion, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        transaccion.delete()
        messages.success(request, 'Transacción eliminada exitosamente.')
        return redirect('transacciones_lista')
    
    context = {
        'transaccion': transaccion,
    }
    return render(request, 'transacciones/eliminar.html', context)

# Vistas para Categorías
@login_required(login_url='/finanzas/')
def categorias_lista(request):
    categorias = Categoria.objects.filter(usuario=request.user).order_by('nombre')
    context = {
        'categorias': categorias,
    }
    return render(request, 'categorias/lista.html', context)

@login_required(login_url='/finanzas/')
def categoria_crear(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría creada exitosamente.')
            return redirect('categorias_lista')
    else:
        form = CategoriaForm(user=request.user)
    
    context = {
        'form': form,
        'titulo': 'Nueva Categoría',
    }
    return render(request, 'categorias/form.html', context)

@login_required(login_url='/finanzas/')
def categoria_editar(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Categoría actualizada exitosamente.')
            return redirect('categorias_lista')
    else:
        form = CategoriaForm(instance=categoria, user=request.user)
    
    context = {
        'form': form,
        'categoria': categoria,
        'titulo': 'Editar Categoría',
    }
    return render(request, 'categorias/form.html', context)

@login_required(login_url='/finanzas/')
def categoria_eliminar(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        categoria.delete()
        messages.success(request, 'Categoría eliminada exitosamente.')
        return redirect('categorias_lista')
    
    context = {
        'categoria': categoria
    }
    return render(request, 'categorias/eliminar.html', context)

# Vistas para Cuentas
@login_required(login_url='/finanzas/')
def cuentas_lista(request):
    cuentas = Cuenta.objects.filter(usuario=request.user).order_by('nombre')
    context = {
        'cuentas': cuentas,
    }
    return render(request, 'cuentas/lista.html', context)

@login_required(login_url='/finanzas/')
def cuenta_crear(request):
    if request.method == 'POST':
        form = CuentaForm(request.POST, user=request.user)
        if form.is_valid():
            cuenta = form.save(commit=False)
            cuenta.saldo_actual = cuenta.saldo_inicial
            cuenta.save()
            messages.success(request, 'Cuenta creada exitosamente.')
            return redirect('cuentas_lista')
    else:
        form = CuentaForm(user=request.user)
    
    context = {
        'form': form,
        'titulo': 'Nueva Cuenta',
    }
    return render(request, 'cuentas/form.html', context)

@login_required(login_url='/finanzas/')
def cuenta_editar(request, pk):
    cuenta = get_object_or_404(Cuenta, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        form = CuentaForm(request.POST, instance=cuenta, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Cuenta actualizada exitosamente.')
            return redirect('cuentas_lista')
    else:
        form = CuentaForm(instance=cuenta, user=request.user)
    
    context = {
        'form': form,
        'cuenta': cuenta,
        'titulo': 'Editar Cuenta',
    }
    return render(request, 'cuentas/form.html', context)

@login_required(login_url='/finanzas/')
def cuenta_eliminar(request, pk):
    cuenta = get_object_or_404(Cuenta, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        cuenta.delete()
        messages.success(request, 'Cuenta eliminada exitosamente.')
        return redirect('cuentas_lista')
    
    context = {
        'cuenta': cuenta
    }
    return render(request, 'cuentas/eliminar.html', context)

# API para obtener datos del dashboard
@login_required(login_url='/finanzas/')
def api_dashboard_data(request):
    """API para obtener datos del dashboard en formato JSON"""
    mes_actual = date.today().replace(day=1)
    mes_siguiente = (mes_actual + timedelta(days=32)).replace(day=1)
    
    transacciones_mes = Transaccion.objects.filter(
        usuario=request.user,
        fecha__gte=mes_actual,
        fecha__lt=mes_siguiente
    )
    
    ingresos_mes = transacciones_mes.filter(tipo='ingreso').aggregate(
        total=Sum('monto'))['total'] or Decimal('0')
    gastos_mes = transacciones_mes.filter(tipo='gasto').aggregate(
        total=Sum('monto'))['total'] or Decimal('0')
    
    # Gastos por categoría para gráfico
    gastos_por_categoria = list(transacciones_mes.filter(tipo='gasto').values(
        'categoria__nombre', 'categoria__color'
    ).annotate(total=Sum('monto')).order_by('-total')[:10])
    
    data = {
        'ingresos_mes': float(ingresos_mes),
        'gastos_mes': float(gastos_mes),
        'balance_mes': float(ingresos_mes - gastos_mes),
        'gastos_por_categoria': gastos_por_categoria,
    }
    
    return JsonResponse(data)

# Vistas para Presupuestos
@login_required
def presupuestos_lista(request):
    presupuestos = Presupuesto.objects.filter(usuario=request.user).order_by('-fecha_creacion')
    
    # Actualizar gastos de todos los presupuestos
    for presupuesto in presupuestos:
        presupuesto.actualizar_gasto()
    
    context = {
        'presupuestos': presupuestos,
        'titulo': 'Presupuestos'
    }
    return render(request, 'presupuestos/lista.html', context)

@login_required
def presupuesto_crear(request):
    if request.method == 'POST':
        form = PresupuestoForm(request.POST, user=request.user)
        if form.is_valid():
            presupuesto = form.save(commit=False)
            presupuesto.usuario = request.user
            presupuesto.save()
            form.save_m2m()  # Guardar las relaciones many-to-many
            messages.success(request, 'Presupuesto creado exitosamente.')
            return redirect('presupuestos_lista')
    else:
        form = PresupuestoForm(user=request.user)
    
    context = {
        'form': form,
        'titulo': 'Crear Presupuesto'
    }
    return render(request, 'presupuestos/form.html', context)

@login_required
def presupuesto_editar(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        form = PresupuestoForm(request.POST, instance=presupuesto, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Presupuesto actualizado exitosamente.')
            return redirect('presupuestos_lista')
    else:
        form = PresupuestoForm(instance=presupuesto, user=request.user)
    
    context = {
        'form': form,
        'presupuesto': presupuesto,
        'titulo': 'Editar Presupuesto'
    }
    return render(request, 'presupuestos/form.html', context)

@login_required
def presupuesto_eliminar(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        presupuesto.delete()
        messages.success(request, 'Presupuesto eliminado exitosamente.')
        return redirect('presupuestos_lista')
    
    context = {
        'presupuesto': presupuesto
    }
    return render(request, 'presupuestos/eliminar.html', context)

# Vistas para Metas
@login_required
def metas_lista(request):
    metas = Meta.objects.filter(usuario=request.user).order_by('-fecha_creacion')
    
    context = {
        'metas': metas,
        'titulo': 'Metas'
    }
    return render(request, 'metas/lista.html', context)

@login_required
def meta_crear(request):
    if request.method == 'POST':
        form = MetaForm(request.POST, user=request.user)
        if form.is_valid():
            meta = form.save(commit=False)
            meta.usuario = request.user
            meta.save()
            messages.success(request, 'Meta creada exitosamente.')
            return redirect('metas_lista')
    else:
        form = MetaForm(user=request.user)
    
    context = {
        'form': form,
        'titulo': 'Crear Meta'
    }
    return render(request, 'metas/form.html', context)

@login_required
def meta_editar(request, pk):
    meta = get_object_or_404(Meta, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        form = MetaForm(request.POST, instance=meta, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Meta actualizada exitosamente.')
            return redirect('metas_lista')
    else:
        form = MetaForm(instance=meta, user=request.user)
    
    context = {
        'form': form,
        'meta': meta,
        'titulo': 'Editar Meta'
    }
    return render(request, 'metas/form.html', context)

@login_required
def meta_eliminar(request, pk):
    meta = get_object_or_404(Meta, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        meta.delete()
        messages.success(request, 'Meta eliminada exitosamente.')
        return redirect('metas_lista')
    
    context = {
        'meta': meta
    }
    return render(request, 'metas/eliminar.html', context)

@login_required
def meta_actualizar_progreso(request, pk):
    meta = get_object_or_404(Meta, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        nuevo_monto = request.POST.get('monto_actual')
        try:
            nuevo_monto = float(nuevo_monto)
            if nuevo_monto >= 0:
                meta.monto_actual = nuevo_monto
                meta.save()
                messages.success(request, 'Progreso de la meta actualizado exitosamente.')
            else:
                messages.error(request, 'El monto debe ser mayor o igual a 0.')
        except ValueError:
            messages.error(request, 'Por favor ingrese un monto válido.')
    
    return redirect('metas_lista')

# Vistas para Tags
@login_required
def tags_lista(request):
    tags = Tag.objects.filter(usuario=request.user).order_by('nombre')
    context = {
        'tags': tags,
    }
    return render(request, 'tags/lista.html', context)

@login_required
def tag_crear(request):
    if request.method == 'POST':
        form = TagForm(request.POST, user=request.user)
        if form.is_valid():
            tag = form.save(commit=False)
            tag.usuario = request.user
            tag.save()
            messages.success(request, 'Tag creado exitosamente.')
            return redirect('tags_lista')
    else:
        form = TagForm(user=request.user)
    
    context = {
        'form': form,
        'titulo': 'Nuevo Tag',
    }
    return render(request, 'tags/form.html', context)

@login_required
def tag_editar(request, pk):
    tag = get_object_or_404(Tag, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        form = TagForm(request.POST, instance=tag, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tag actualizado exitosamente.')
            return redirect('tags_lista')
    else:
        form = TagForm(instance=tag, user=request.user)
    
    context = {
        'form': form,
        'tag': tag,
        'titulo': 'Editar Tag',
    }
    return render(request, 'tags/form.html', context)

@login_required
def tag_eliminar(request, pk):
    tag = get_object_or_404(Tag, pk=pk, usuario=request.user)
    
    if request.method == 'POST':
        tag.delete()
        messages.success(request, 'Tag eliminado exitosamente.')
        return redirect('tags_lista')
    
    context = {
        'tag': tag
    }
    return render(request, 'tags/eliminar.html', context)

@login_required
def corte_mes_confirmar(request):
    """Vista para confirmar el corte de mes"""
    # Obtener el mes actual
    hoy = date.today()
    mes_actual = hoy.month
    año_actual = hoy.year
    
    # Verificar si ya se hizo el corte para este mes
    corte_existente = CorteMes.objects.filter(
        usuario=request.user,
        mes_cortado=mes_actual,
        año_cortado=año_actual
    ).first()
    
    if corte_existente:
        messages.warning(request, f'Ya se realizó el corte para {corte_existente.mes_nombre} {año_actual}.')
        return redirect('dashboard')
    
    # Calcular el último día del mes
    ultimo_dia = monthrange(año_actual, mes_actual)[1]
    fecha_fin_mes = date(año_actual, mes_actual, ultimo_dia)
    
    # Calcular estadísticas del mes
    inicio_mes = date(año_actual, mes_actual, 1)
    
    ingresos_mes = Transaccion.objects.filter(
        usuario=request.user,
        tipo='ingreso',
        fecha__gte=inicio_mes,
        fecha__lte=fecha_fin_mes
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
    
    gastos_mes = Transaccion.objects.filter(
        usuario=request.user,
        tipo='gasto',
        fecha__gte=inicio_mes,
        fecha__lte=fecha_fin_mes
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
    
    balance_mes = ingresos_mes - gastos_mes
    
    # Obtener saldos actuales de las cuentas
    cuentas_activas = Cuenta.objects.filter(usuario=request.user, activa=True)
    saldos_cuentas = {}
    
    for cuenta in cuentas_activas:
        saldos_cuentas[cuenta.id] = {
            'nombre': cuenta.nombre,
            'saldo_actual': float(cuenta.saldo_actual),
            'tipo': cuenta.tipo_cuenta
        }
    
    # Transacciones del mes
    transacciones_mes = Transaccion.objects.filter(
        usuario=request.user,
        fecha__gte=inicio_mes,
        fecha__lte=fecha_fin_mes
    ).count()
    
    # CALCULAR ESTADÍSTICAS DE GASTOS COMPARTIDOS
    # Gastos compartidos del mes donde el usuario es miembro
    grupos_usuario = GrupoGastosCompartidos.objects.filter(
        miembros=request.user,
        activo=True
    )
    
    gastos_compartidos_mes = GastoCompartido.objects.filter(
        grupo__in=grupos_usuario,
        fecha__gte=inicio_mes,
        fecha__lte=fecha_fin_mes,
        activo=True
    )
    
    total_gastos_compartidos_mes = gastos_compartidos_mes.aggregate(
        total=Sum('monto_total')
    )['total'] or Decimal('0')
    
    # Gastos compartidos pagados por el usuario en el mes
    gastos_pagados_por_usuario = gastos_compartidos_mes.filter(
        pagado_por=request.user
    ).aggregate(total=Sum('monto_total'))['total'] or Decimal('0')
    
    # Pagos realizados por el usuario en gastos compartidos del mes
    pagos_usuario_mes = PagoGastoCompartido.objects.filter(
        miembro=request.user,
        gasto_compartido__fecha__gte=inicio_mes,
        gasto_compartido__fecha__lte=fecha_fin_mes,
        gasto_compartido__activo=True
    ).aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0')
    
    # Saldos pendientes en gastos compartidos al final del mes
    saldos_pendientes_gastos_compartidos = {}
    for grupo in grupos_usuario:
        gastos_grupo = GastoCompartido.objects.filter(
            grupo=grupo,
            activo=True
        )
        
        total_debido_usuario = gastos_grupo.aggregate(
            total=Sum('monto_total')
        )['total'] or Decimal('0')
        
        # Calcular cuánto debe pagar el usuario (dividir por cantidad de miembros)
        cantidad_miembros = grupo.cantidad_miembros
        if cantidad_miembros > 0:
            monto_por_persona = total_debido_usuario / cantidad_miembros
        else:
            monto_por_persona = Decimal('0')
        
        # Calcular cuánto ya pagó el usuario
        pagos_usuario_grupo = PagoGastoCompartido.objects.filter(
            miembro=request.user,
            gasto_compartido__grupo=grupo,
            gasto_compartido__activo=True
        ).aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0')
        
        saldo_pendiente = monto_por_persona - pagos_usuario_grupo
        
        saldos_pendientes_gastos_compartidos[grupo.id] = {
            'nombre': grupo.nombre,
            'total_debido': float(monto_por_persona),
            'total_pagado': float(pagos_usuario_grupo),
            'saldo_pendiente': float(saldo_pendiente),
            'cantidad_miembros': cantidad_miembros
        }
    
    # Total de saldos pendientes en gastos compartidos
    total_saldo_pendiente_gastos_compartidos = sum(
        saldo['saldo_pendiente'] for saldo in saldos_pendientes_gastos_compartidos.values()
    )
    
    context = {
        'mes_actual': mes_actual,
        'año_actual': año_actual,
        'mes_nombre': CorteMes(mes_cortado=mes_actual).mes_nombre,
        'fecha_fin_mes': fecha_fin_mes,
        'ingresos_mes': ingresos_mes,
        'gastos_mes': gastos_mes,
        'balance_mes': balance_mes,
        'saldos_cuentas': saldos_cuentas,
        'transacciones_mes': transacciones_mes,
        'cuentas_activas': cuentas_activas,
        # Nuevos datos de gastos compartidos
        'total_gastos_compartidos_mes': total_gastos_compartidos_mes,
        'gastos_pagados_por_usuario': gastos_pagados_por_usuario,
        'pagos_usuario_mes': pagos_usuario_mes,
        'saldos_pendientes_gastos_compartidos': saldos_pendientes_gastos_compartidos,
        'total_saldo_pendiente_gastos_compartidos': total_saldo_pendiente_gastos_compartidos,
        'grupos_usuario': grupos_usuario,
    }
    
    return render(request, 'corte_mes/confirmar.html', context)

@login_required
def corte_mes_ejecutar(request):
    """Vista para ejecutar el corte de mes"""
    if request.method != 'POST':
        return redirect('corte_mes_confirmar')
    
    # Obtener el mes actual
    hoy = date.today()
    mes_actual = hoy.month
    año_actual = hoy.year
    
    # Verificar si ya se hizo el corte para este mes
    corte_existente = CorteMes.objects.filter(
        usuario=request.user,
        mes_cortado=mes_actual,
        año_cortado=año_actual
    ).first()
    
    if corte_existente:
        messages.warning(request, f'Ya se realizó el corte para {corte_existente.mes_nombre} {año_actual}.')
        return redirect('dashboard')
    
    # Calcular el último día del mes
    ultimo_dia = monthrange(año_actual, mes_actual)[1]
    fecha_fin_mes = date(año_actual, mes_actual, ultimo_dia)
    
    # Calcular estadísticas del mes
    inicio_mes = date(año_actual, mes_actual, 1)
    
    ingresos_mes = Transaccion.objects.filter(
        usuario=request.user,
        tipo='ingreso',
        fecha__gte=inicio_mes,
        fecha__lte=fecha_fin_mes
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
    
    gastos_mes = Transaccion.objects.filter(
        usuario=request.user,
        tipo='gasto',
        fecha__gte=inicio_mes,
        fecha__lte=fecha_fin_mes
    ).aggregate(total=Sum('monto'))['total'] or Decimal('0')
    
    balance_mes = ingresos_mes - gastos_mes
    
    # Obtener saldos actuales de las cuentas
    cuentas_activas = Cuenta.objects.filter(usuario=request.user, activa=True)
    saldos_cuentas = {}
    
    for cuenta in cuentas_activas:
        saldos_cuentas[cuenta.id] = float(cuenta.saldo_actual)
    
    # CALCULAR ESTADÍSTICAS DE GASTOS COMPARTIDOS
    grupos_usuario = GrupoGastosCompartidos.objects.filter(
        miembros=request.user,
        activo=True
    )
    
    gastos_compartidos_mes = GastoCompartido.objects.filter(
        grupo__in=grupos_usuario,
        fecha__gte=inicio_mes,
        fecha__lte=fecha_fin_mes,
        activo=True
    )
    
    total_gastos_compartidos_mes = gastos_compartidos_mes.aggregate(
        total=Sum('monto_total')
    )['total'] or Decimal('0')
    
    # Gastos compartidos pagados por el usuario en el mes
    gastos_pagados_por_usuario = gastos_compartidos_mes.filter(
        pagado_por=request.user
    ).aggregate(total=Sum('monto_total'))['total'] or Decimal('0')
    
    # Pagos realizados por el usuario en gastos compartidos del mes
    pagos_usuario_mes = PagoGastoCompartido.objects.filter(
        miembro=request.user,
        gasto_compartido__fecha__gte=inicio_mes,
        gasto_compartido__fecha__lte=fecha_fin_mes,
        gasto_compartido__activo=True
    ).aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0')
    
    # Saldos pendientes en gastos compartidos al final del mes
    saldos_pendientes_gastos_compartidos = {}
    for grupo in grupos_usuario:
        gastos_grupo = GastoCompartido.objects.filter(
            grupo=grupo,
            activo=True
        )
        
        total_debido_usuario = gastos_grupo.aggregate(
            total=Sum('monto_total')
        )['total'] or Decimal('0')
        
        # Calcular cuánto debe pagar el usuario (dividir por cantidad de miembros)
        cantidad_miembros = grupo.cantidad_miembros
        if cantidad_miembros > 0:
            monto_por_persona = total_debido_usuario / cantidad_miembros
        else:
            monto_por_persona = Decimal('0')
        
        # Calcular cuánto ya pagó el usuario
        pagos_usuario_grupo = PagoGastoCompartido.objects.filter(
            miembro=request.user,
            gasto_compartido__grupo=grupo,
            gasto_compartido__activo=True
        ).aggregate(total=Sum('monto_pagado'))['total'] or Decimal('0')
        
        saldo_pendiente = monto_por_persona - pagos_usuario_grupo
        
        saldos_pendientes_gastos_compartidos[grupo.id] = {
            'nombre': grupo.nombre,
            'total_debido': float(monto_por_persona),
            'total_pagado': float(pagos_usuario_grupo),
            'saldo_pendiente': float(saldo_pendiente),
            'cantidad_miembros': cantidad_miembros
        }
    
    # Total de saldos pendientes en gastos compartidos
    total_saldo_pendiente_gastos_compartidos = sum(
        saldo['saldo_pendiente'] for saldo in saldos_pendientes_gastos_compartidos.values()
    )
    
    # Crear el registro del corte con datos de gastos compartidos
    corte_mes = CorteMes.objects.create(
        usuario=request.user,
        fecha_corte=fecha_fin_mes,
        mes_cortado=mes_actual,
        año_cortado=año_actual,
        total_ingresos=ingresos_mes,
        total_gastos=gastos_mes,
        balance_mes=balance_mes,
        saldos_cuentas=saldos_cuentas,
        mantener_saldos=True
    )
    
    # Guardar datos de gastos compartidos en el campo saldos_cuentas (extendemos su uso)
    datos_completos = {
        'saldos_cuentas': saldos_cuentas,
        'gastos_compartidos': {
            'total_gastos_compartidos_mes': float(total_gastos_compartidos_mes),
            'gastos_pagados_por_usuario': float(gastos_pagados_por_usuario),
            'pagos_usuario_mes': float(pagos_usuario_mes),
            'saldos_pendientes_grupos': saldos_pendientes_gastos_compartidos,
            'total_saldo_pendiente': float(total_saldo_pendiente_gastos_compartidos)
        }
    }
    
    # Actualizar el campo saldos_cuentas con todos los datos
    corte_mes.saldos_cuentas = datos_completos
    corte_mes.save()
    
    # Opcional: Archivar transacciones del mes (mover a una tabla de histórico)
    # Por ahora solo registramos el corte sin eliminar transacciones
    
    messages.success(
        request, 
        f'Corte de mes realizado exitosamente para {corte_mes.mes_nombre} {año_actual}. '
        f'Balance: ${balance_mes:,.2f} | Gastos Compartidos: ${total_gastos_compartidos_mes:,.2f}'
    )
    
    return redirect('dashboard')

@login_required
def cortes_mes_lista(request):
    """Vista para listar todos los cortes de mes"""
    cortes = CorteMes.objects.filter(usuario=request.user).order_by('-fecha_corte')
    
    # Paginación
    paginator = Paginator(cortes, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_cortes': cortes.count(),
    }
    
    return render(request, 'corte_mes/lista.html', context)

@login_required
def corte_mes_detalle(request, pk):
    """Vista para mostrar el detalle de un corte de mes"""
    corte = get_object_or_404(CorteMes, pk=pk, usuario=request.user)
    
    # Obtener las transacciones del mes cortado
    inicio_mes = date(corte.año_cortado, corte.mes_cortado, 1)
    ultimo_dia = monthrange(corte.año_cortado, corte.mes_cortado)[1]
    fecha_fin_mes = date(corte.año_cortado, corte.mes_cortado, ultimo_dia)
    
    transacciones = Transaccion.objects.filter(
        usuario=request.user,
        fecha__gte=inicio_mes,
        fecha__lte=fecha_fin_mes
    ).select_related('categoria', 'cuenta').order_by('-fecha')
    
    # Gastos por categoría
    gastos_por_categoria = Transaccion.objects.filter(
        usuario=request.user,
        tipo='gasto',
        fecha__gte=inicio_mes,
        fecha__lte=fecha_fin_mes
    ).values('categoria__nombre', 'categoria__color').annotate(
        total=Sum('monto')
    ).order_by('-total')
    
    # Extraer datos de gastos compartidos del corte
    datos_gastos_compartidos = {}
    saldos_cuentas_originales = {}
    
    if isinstance(corte.saldos_cuentas, dict):
        if 'gastos_compartidos' in corte.saldos_cuentas:
            # Formato nuevo con gastos compartidos
            datos_gastos_compartidos = corte.saldos_cuentas.get('gastos_compartidos', {})
            saldos_cuentas_originales = corte.saldos_cuentas.get('saldos_cuentas', {})
        else:
            # Formato antiguo, solo saldos de cuentas
            saldos_cuentas_originales = corte.saldos_cuentas
    
    # Obtener información actualizada de gastos compartidos del mes
    grupos_usuario = GrupoGastosCompartidos.objects.filter(
        miembros=request.user,
        activo=True
    )
    
    gastos_compartidos_mes = GastoCompartido.objects.filter(
        grupo__in=grupos_usuario,
        fecha__gte=inicio_mes,
        fecha__lte=fecha_fin_mes,
        activo=True
    ).select_related('grupo', 'pagado_por').order_by('-fecha')
    
    # Pagos del usuario en gastos compartidos del mes
    pagos_usuario_mes = PagoGastoCompartido.objects.filter(
        miembro=request.user,
        gasto_compartido__fecha__gte=inicio_mes,
        gasto_compartido__fecha__lte=fecha_fin_mes,
        gasto_compartido__activo=True
    ).select_related('gasto_compartido', 'gasto_compartido__grupo').order_by('-fecha_pago')
    
    context = {
        'corte': corte,
        'transacciones': transacciones,
        'gastos_por_categoria': gastos_por_categoria,
        'total_transacciones': transacciones.count(),
        # Datos de gastos compartidos
        'datos_gastos_compartidos': datos_gastos_compartidos,
        'saldos_cuentas_originales': saldos_cuentas_originales,
        'grupos_usuario': grupos_usuario,
        'gastos_compartidos_mes': gastos_compartidos_mes,
        'pagos_usuario_mes': pagos_usuario_mes,
        'total_gastos_compartidos': gastos_compartidos_mes.count(),
        'total_pagos_usuario': pagos_usuario_mes.count(),
    }
    
    return render(request, 'corte_mes/detalle.html', context)

@login_required
def usuarios_lista(request):
    usuarios = User.objects.filter(is_superuser=False).order_by('username')
    context = {
        'usuarios': usuarios,
    }
    return render(request, 'usuarios/lista.html', context)

@login_required
def usuario_crear(request):
    if request.method == 'POST':
        form = UsuarioCrearForm(request.POST, user=request.user)
        if form.is_valid():
            usuario = form.save(commit=False)
            usuario.password = make_password(form.cleaned_data['password'])
            usuario.save()
            messages.success(request, 'Usuario creado exitosamente.')
            return redirect('usuarios_lista')
    else:
        form = UsuarioCrearForm(user=request.user)
    
    context = {
        'form': form,
        'titulo': 'Nuevo Usuario',
    }
    return render(request, 'usuarios/form.html', context)

@login_required
def usuario_editar(request, pk):
    usuario = get_object_or_404(User, pk=pk, is_superuser=False)
    
    if request.method == 'POST':
        form = UsuarioEditarForm(request.POST, instance=usuario, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario actualizado exitosamente.')
            return redirect('usuarios_lista')
    else:
        form = UsuarioEditarForm(instance=usuario, user=request.user)
    
    context = {
        'form': form,
        'usuario': usuario,
        'titulo': 'Editar Usuario',
    }
    return render(request, 'usuarios/form.html', context)

@login_required
def usuario_eliminar(request, pk):
    usuario = get_object_or_404(User, pk=pk, is_superuser=False)
    
    if request.method == 'POST':
        usuario.delete()
        messages.success(request, 'Usuario eliminado exitosamente.')
        return redirect('usuarios_lista')
    
    context = {
        'usuario': usuario
    }
    return render(request, 'usuarios/eliminar.html', context)

@login_required
def configuracion_usuario(request):
    if request.method == 'POST':
        form = ConfiguracionUsuarioForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuración del usuario actualizada exitosamente.')
            return redirect('dashboard')
    else:
        form = ConfiguracionUsuarioForm(instance=request.user)
    
    context = {
        'form': form,
        'titulo': 'Configuración del Usuario',
    }
    return render(request, 'configuracion_usuario/form.html', context)

@login_required
def configuracion_sistema(request):
    if request.method == 'POST':
        form = ConfiguracionSistemaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Configuración del sistema actualizada exitosamente.')
            return redirect('dashboard')
    else:
        form = ConfiguracionSistemaForm()
    
    context = {
        'form': form,
        'titulo': 'Configuración del Sistema',
    }
    return render(request, 'configuracion_sistema/form.html', context)

# Vistas para Configuración del Sistema
@login_required
@staff_required
def configuracion(request):
    """Vista principal de configuración del sistema"""
    # Obtener todos los usuarios
    usuarios = User.objects.all().order_by('username')
    
    # Obtener configuración del sistema (usar la del usuario actual como ejemplo)
    try:
        config = ConfiguracionUsuario.objects.get(usuario=request.user)
    except ConfiguracionUsuario.DoesNotExist:
        config = ConfiguracionUsuario.objects.create(usuario=request.user)
    
    # Estadísticas del sistema
    total_usuarios = User.objects.count()
    usuarios_pendientes = User.objects.filter(is_active=False).count()
    total_transacciones = Transaccion.objects.count()
    total_cuentas = Cuenta.objects.count()
    total_categorias = Categoria.objects.count()
    
    context = {
        'usuarios': usuarios,
        'config': config,
        'total_usuarios': total_usuarios,
        'usuarios_pendientes': usuarios_pendientes,
        'total_transacciones': total_transacciones,
        'total_cuentas': total_cuentas,
        'total_categorias': total_categorias,
    }
    return render(request, 'configuracion/configuracion.html', context)

@login_required
@staff_required
def usuarios_pendientes(request):
    """Vista para listar usuarios pendientes de activación"""
    usuarios_pendientes = User.objects.filter(is_active=False).order_by('date_joined')
    
    context = {
        'usuarios_pendientes': usuarios_pendientes,
        'total_pendientes': usuarios_pendientes.count(),
    }
    return render(request, 'configuracion/usuarios_pendientes.html', context)

@login_required
@staff_required
def usuario_crear(request):
    """Crear nuevo usuario"""
    if request.method == 'POST':
        form = UsuarioCrearForm(request.POST)
        config_form = ConfiguracionUsuarioForm(request.POST)
        
        if form.is_valid() and config_form.is_valid():
            usuario = form.save()
            
            # Crear configuración para el usuario
            config = config_form.save(commit=False)
            config.usuario = usuario
            config.save()
            
            messages.success(request, f'Usuario "{usuario.username}" creado exitosamente.')
            return redirect('configuracion')
    else:
        form = UsuarioCrearForm()
        config_form = ConfiguracionUsuarioForm()
    
    context = {
        'form': form,
        'config_form': config_form,
        'titulo': 'Nuevo Usuario',
    }
    return render(request, 'configuracion/usuario_form.html', context)

@login_required
@staff_required
def usuario_editar(request, pk):
    """Editar usuario existente"""
    usuario = get_object_or_404(User, pk=pk)
    
    # Obtener o crear configuración del usuario
    try:
        config = ConfiguracionUsuario.objects.get(usuario=usuario)
    except ConfiguracionUsuario.DoesNotExist:
        config = ConfiguracionUsuario.objects.create(usuario=usuario)
    
    if request.method == 'POST':
        form = UsuarioEditarForm(request.POST, instance=usuario)
        config_form = ConfiguracionUsuarioForm(request.POST, instance=config)
        
        if form.is_valid() and config_form.is_valid():
            # Manejar cambio de contraseña
            new_password = request.POST.get('new_password')
            if new_password:
                usuario.password = make_password(new_password)
            
            form.save()
            config_form.save()
            
            messages.success(request, f'Usuario "{usuario.username}" actualizado exitosamente.')
            return redirect('configuracion')
    else:
        form = UsuarioEditarForm(instance=usuario)
        config_form = ConfiguracionUsuarioForm(instance=config)
    
    context = {
        'form': form,
        'config_form': config_form,
        'usuario': usuario,
        'titulo': 'Editar Usuario',
    }
    return render(request, 'configuracion/usuario_form.html', context)

@login_required
@staff_required
def usuario_eliminar(request, pk):
    """Eliminar usuario"""
    usuario = get_object_or_404(User, pk=pk)
    
    # No permitir eliminar superusuarios
    if usuario.is_superuser:
        messages.error(request, 'No se puede eliminar un superusuario.')
        return redirect('configuracion')
    
    if request.method == 'POST':
        username = usuario.username
        usuario.delete()
        messages.success(request, f'Usuario "{username}" eliminado exitosamente.')
        
        # Redirigir a usuarios_pendientes si viene de esa página, sino a configuracion
        referer = request.META.get('HTTP_REFERER', '')
        if 'usuarios-pendientes' in referer:
            return redirect('usuarios_pendientes')
        return redirect('configuracion')
    
    context = {
        'usuario': usuario
    }
    return render(request, 'configuracion/usuario_eliminar.html', context)

@login_required
@staff_required
def usuario_activar(request, pk):
    """Activar usuario"""
    usuario = get_object_or_404(User, pk=pk)
    usuario.is_active = True
    usuario.save()
    messages.success(request, f'Usuario "{usuario.username}" activado exitosamente.')
    
    # Redirigir a usuarios_pendientes si viene de esa página, sino a configuracion
    referer = request.META.get('HTTP_REFERER', '')
    if 'usuarios-pendientes' in referer:
        return redirect('usuarios_pendientes')
    return redirect('configuracion')

@login_required
@staff_required
def usuario_desactivar(request, pk):
    """Desactivar usuario"""
    usuario = get_object_or_404(User, pk=pk)
    
    # No permitir desactivar superusuarios
    if usuario.is_superuser:
        messages.error(request, 'No se puede desactivar un superusuario.')
        return redirect('configuracion')
    
    usuario.is_active = False
    usuario.save()
    messages.success(request, f'Usuario "{usuario.username}" desactivado exitosamente.')
    return redirect('configuracion')

@login_required
@staff_required
def configuracion_guardar(request):
    """Guardar configuración del sistema"""
    if request.method == 'POST':
        # Obtener configuración del usuario actual
        try:
            config = ConfiguracionUsuario.objects.get(usuario=request.user)
        except ConfiguracionUsuario.DoesNotExist:
            config = ConfiguracionUsuario.objects.create(usuario=request.user)
        
        # Actualizar configuración
        config.moneda_principal = request.POST.get('moneda_principal', 'ARS')
        config.zona_horaria = request.POST.get('zona_horaria', 'America/Argentina/Buenos_Aires')
        config.notificaciones_activas = request.POST.get('notificaciones_activas') == 'on'
        config.recordatorios_pago = request.POST.get('recordatorios_pago') == 'on'
        config.save()
        
        messages.success(request, 'Configuración guardada exitosamente.')
    
    return redirect('configuracion')

# Vistas para acciones del sistema
@login_required
@staff_required
def backup_datos(request):
    """Crear backup de datos"""
    # Aquí implementarías la lógica de backup
    messages.info(request, 'Función de backup en desarrollo.')
    return redirect('configuracion')

@login_required
@staff_required
def limpiar_datos(request):
    """Limpiar datos antiguos"""
    # Aquí implementarías la lógica de limpieza
    messages.info(request, 'Función de limpieza en desarrollo.')
    return redirect('configuracion')

@login_required
@staff_required
def exportar_datos(request):
    """Exportar datos"""
    # Aquí implementarías la lógica de exportación
    messages.info(request, 'Función de exportación en desarrollo.')
    return redirect('configuracion')

@login_required
@staff_required
def importar_datos(request):
    """Importar datos"""
    # Aquí implementarías la lógica de importación
    messages.info(request, 'Función de importación en desarrollo.')
    return redirect('configuracion')

@login_required
def configuracion_personal(request):
    """Vista para configuración personal del usuario"""
    # Obtener o crear configuración del usuario
    try:
        config = ConfiguracionUsuario.objects.get(usuario=request.user)
    except ConfiguracionUsuario.DoesNotExist:
        config = ConfiguracionUsuario.objects.create(usuario=request.user)
    
    if request.method == 'POST':
        form = ConfiguracionUsuarioForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tu configuración personal ha sido actualizada exitosamente.')
            return redirect('dashboard')
    else:
        form = ConfiguracionUsuarioForm(instance=config)
    
    context = {
        'form': form,
        'config': config,
        'titulo': 'Mi Configuración',
    }
    return render(request, 'configuracion/configuracion_personal.html', context)

# Vistas para Gastos Compartidos
@login_required
def dashboard_gastos_compartidos(request):
    """Dashboard principal de gastos compartidos mejorado"""
    from datetime import date, timedelta
    try:
        # Obtener grupos donde el usuario es miembro
        grupos = GrupoGastosCompartidos.objects.filter(miembros=request.user, activo=True)
        
        # Estadísticas generales
        total_grupos = grupos.count()
        total_gastos = GastoCompartido.objects.filter(grupo__in=grupos, activo=True).count()
        
        # Calcular total de gastos del usuario
        gastos_usuario = GastoCompartido.objects.filter(
            grupo__in=grupos, 
            activo=True
        ).select_related('grupo')
        
        total_gastado = sum(gasto.monto_por_persona for gasto in gastos_usuario)
        
        # Calcular pagos pendientes del usuario
        pagos_pendientes = PagoGastoCompartido.objects.filter(
            miembro=request.user,
            gasto_compartido__activo=True,
            estado='pendiente'
        ).select_related('gasto_compartido', 'gasto_compartido__grupo')
        
        total_pendiente = sum(pago.monto_debido - pago.monto_pagado for pago in pagos_pendientes)
        
        # Obtener gastos recientes (últimos 5, incluyendo cancelados)
        gastos_recientes = GastoCompartido.objects.filter(
            grupo__in=grupos, 
            activo=True
        ).select_related('grupo', 'pagado_por').order_by('-fecha')[:5]
        
        # Calcular estadísticas adicionales
        gastos_pendientes = GastoCompartido.objects.filter(
            grupo__in=grupos, 
            activo=True,
            estado='pendiente'
        ).count()
        
        gastos_vencidos = GastoCompartido.objects.filter(
            grupo__in=grupos, 
            activo=True,
            estado='vencido'
        ).count()
        
        gastos_cancelados = GastoCompartido.objects.filter(
            grupo__in=grupos,
            activo=True,
            estado='cancelado'
        ).count()
        
        # Gastos próximos a vencer (en los próximos 3 días)
        hoy = date.today()
        proximos_vencer = GastoCompartido.objects.filter(
            grupo__in=grupos,
            activo=True,
            estado='pendiente',
            fecha_vencimiento__isnull=False,
            fecha_vencimiento__gte=hoy,
            fecha_vencimiento__lte=hoy + timedelta(days=3)
        ).order_by('fecha_vencimiento')
        
        # Obtener pagos pendientes del usuario para mostrar en la sección correspondiente
        pagos_pendientes_usuario = PagoGastoCompartido.objects.filter(
            miembro=request.user,
            gasto_compartido__activo=True,
            estado='pendiente'
        ).select_related('gasto_compartido', 'gasto_compartido__grupo')
        
        # Resumen por grupo: saldo neto del usuario en cada grupo
        grupos_con_saldo = []
        for grupo in grupos:
            pagos_miembro = PagoGastoCompartido.objects.filter(
                gasto_compartido__grupo=grupo,
                miembro=request.user
            )
            total_debido = sum(p.monto_debido for p in pagos_miembro)
            total_pagado = sum(p.monto_pagado for p in pagos_miembro)
            saldo_neto = total_debido - total_pagado
            grupos_con_saldo.append({
                'grupo': grupo,
                'saldo_neto': saldo_neto
            })
        
        # Filtros rápidos (opcional: puedes expandir esto en el template)
        estados_posibles = [
            {'valor': 'pendiente', 'nombre': 'Pendiente'},
            {'valor': 'pagado', 'nombre': 'Pagado'},
            {'valor': 'vencido', 'nombre': 'Vencido'},
            {'valor': 'cancelado', 'nombre': 'Cancelado'},
        ]
        
        context = {
            'grupos': grupos,  # Lista directa de grupos
            'grupos_con_saldo': grupos_con_saldo,
            'total_grupos': total_grupos,
            'total_gastos': total_gastos,
            'total_gastado': total_gastado,
            'total_pendiente': total_pendiente,
            'gastos_recientes': gastos_recientes,
            'pagos_pendientes': pagos_pendientes,
            'pagos_pendientes_usuario': pagos_pendientes_usuario,
            'gastos_pendientes': gastos_pendientes,
            'gastos_vencidos': gastos_vencidos,
            'gastos_cancelados': gastos_cancelados,
            'total_adeudado': total_pendiente,  # Alias para el template
            'proximos_vencer': proximos_vencer,
            'estados_posibles': estados_posibles,
        }
        return render(request, 'gastos_compartidos/dashboard.html', context)
    except Exception as e:
        messages.error(request, f'Error al cargar el dashboard: {str(e)}')
        return redirect('dashboard')

@login_required
def grupos_gastos_compartidos_lista(request):
    """Lista de grupos de gastos compartidos del usuario"""
    grupos = GrupoGastosCompartidos.objects.filter(miembros=request.user, activo=True).order_by('nombre')
    
    context = {
        'grupos': grupos,
    }
    return render(request, 'gastos_compartidos/grupos_lista.html', context)

@login_required
def grupo_gastos_compartidos_crear(request):
    """Crear un nuevo grupo de gastos compartidos"""
    if request.method == 'POST':
        form = GrupoGastosCompartidosForm(request.POST, user=request.user)
        if form.is_valid():
            grupo = form.save()
            messages.success(request, 'Grupo de gastos compartidos creado exitosamente.')
            return redirect('grupos_gastos_compartidos_lista')
    else:
        form = GrupoGastosCompartidosForm(user=request.user)
    
    context = {
        'form': form,
        'titulo': 'Nuevo Grupo de Gastos Compartidos',
    }
    return render(request, 'gastos_compartidos/grupo_form.html', context)

@login_required
def grupo_gastos_compartidos_editar(request, pk):
    """Editar un grupo de gastos compartidos"""
    grupo = get_object_or_404(GrupoGastosCompartidos, pk=pk, creador=request.user)
    
    if request.method == 'POST':
        form = GrupoGastosCompartidosForm(request.POST, instance=grupo, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Grupo actualizado exitosamente.')
            return redirect('grupos_gastos_compartidos_lista')
    else:
        form = GrupoGastosCompartidosForm(instance=grupo, user=request.user)
    
    context = {
        'form': form,
        'grupo': grupo,
        'titulo': 'Editar Grupo de Gastos Compartidos',
    }
    return render(request, 'gastos_compartidos/grupo_form.html', context)

@login_required
def grupo_gastos_compartidos_eliminar(request, pk):
    """Eliminar un grupo de gastos compartidos - Versión simple sin AJAX"""
    try:
        grupo = get_object_or_404(GrupoGastosCompartidos, pk=pk)
        
        # Verificar permisos
        if grupo.creador != request.user:
            messages.error(request, 'Solo el creador puede eliminar el grupo.')
            return redirect('grupos_gastos_compartidos_lista')
        
        # Si es POST, eliminar directamente
        if request.method == 'POST':
            try:
                # Obtener información antes de eliminar para el mensaje
                nombre_grupo = grupo.nombre
                gastos_count = GastoCompartido.objects.filter(grupo=grupo).count()
                
                # Marcar todos los gastos del grupo como inactivos
                GastoCompartido.objects.filter(grupo=grupo).update(activo=False)
                
                # Marcar el grupo como inactivo
                grupo.activo = False
                grupo.save()
                
                messages.success(request, f'Grupo "{nombre_grupo}" eliminado exitosamente junto con {gastos_count} gastos asociados.')
                return redirect('grupos_gastos_compartidos_lista')
                
            except Exception as e:
                messages.error(request, f'Error al eliminar el grupo: {str(e)}')
                return redirect('grupos_gastos_compartidos_lista')
        
        # Si es GET, mostrar confirmación
        gastos_count = GastoCompartido.objects.filter(grupo=grupo).count()
        pagos_count = PagoGastoCompartido.objects.filter(gasto_compartido__grupo=grupo).count()
        
        context = {
            'grupo': grupo,
            'gastos_count': gastos_count,
            'pagos_count': pagos_count,
        }
        return render(request, 'gastos_compartidos/grupo_eliminar.html', context)
        
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('grupos_gastos_compartidos_lista')

@login_required
def gastos_compartidos_lista(request):
    """Lista de gastos compartidos del usuario"""
    form = FiltroGastosCompartidosForm(request.GET, user=request.user)
    
    # Obtener gastos de grupos donde el usuario es miembro (solo activos)
    gastos = GastoCompartido.objects.filter(grupo__miembros=request.user, activo=True).select_related('grupo', 'pagado_por')
    
    # Aplicar filtros
    if form.is_valid():
        if form.cleaned_data.get('fecha_desde'):
            gastos = gastos.filter(fecha__gte=form.cleaned_data['fecha_desde'])
        if form.cleaned_data.get('fecha_hasta'):
            gastos = gastos.filter(fecha__lte=form.cleaned_data['fecha_hasta'])
        if form.cleaned_data.get('tipo'):
            gastos = gastos.filter(tipo=form.cleaned_data['tipo'])
        if form.cleaned_data.get('estado'):
            gastos = gastos.filter(estado=form.cleaned_data['estado'])
        if form.cleaned_data.get('grupo'):
            gastos = gastos.filter(grupo=form.cleaned_data['grupo'])
        if form.cleaned_data.get('monto_minimo'):
            gastos = gastos.filter(monto_total__gte=form.cleaned_data['monto_minimo'])
        if form.cleaned_data.get('monto_maximo'):
            gastos = gastos.filter(monto_total__lte=form.cleaned_data['monto_maximo'])
    
    # Paginación
    paginator = Paginator(gastos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_gastos': gastos.count(),
    }
    return render(request, 'gastos_compartidos/lista.html', context)

@login_required
def gasto_compartido_crear(request):
    """Crear un nuevo gasto compartido"""
    try:
        grupos = GrupoGastosCompartidos.objects.filter(miembros=request.user, activo=True)
        if not grupos.exists():
            messages.error(request, 'Debe crear un grupo de gastos compartidos antes de agregar gastos.')
            return redirect('grupos_gastos_compartidos_lista')
        if request.method == 'POST':
            form = GastoCompartidoForm(request.POST, request.FILES, user=request.user)
            if form.is_valid():
                gasto = form.save(commit=False)
                grupo = gasto.grupo  # El grupo viene del formulario
                es_en_cuotas = form.cleaned_data.get('es_en_cuotas')
                numero_cuotas = form.cleaned_data.get('numero_cuotas')
                cuota_actual = form.cleaned_data.get('cuota_actual')
                fecha = form.cleaned_data.get('fecha')
                from dateutil.relativedelta import relativedelta
                if es_en_cuotas and numero_cuotas and cuota_actual:
                    for i in range(cuota_actual, numero_cuotas + 1):
                        nueva_fecha = fecha + relativedelta(months=i - cuota_actual)
                        nuevo_gasto = GastoCompartido(
                            grupo=grupo,
                            titulo=f"{gasto.titulo} (Cuota {i}/{numero_cuotas})",
                            descripcion=gasto.descripcion,
                            monto_total=gasto.monto_total,
                            fecha=nueva_fecha,
                            fecha_vencimiento=gasto.fecha_vencimiento,
                            tipo=gasto.tipo,
                            estado=gasto.estado,
                            pagado_por=gasto.pagado_por,
                            cuenta_pago=gasto.cuenta_pago,
                            imagen_recibo=gasto.imagen_recibo,
                            activo=True,
                            es_en_cuotas=True,
                            numero_cuotas=numero_cuotas,
                            cuota_actual=i,
                            fecha_fin_cuotas=gasto.fecha_fin_cuotas
                        )
                        nuevo_gasto.save()
                        # Determinar quién pagó el gasto y cuánto pagó inicialmente
                        pagado_por_usuario = nuevo_gasto.pagado_por
                        monto_pagado_inicial = form.cleaned_data.get('monto_pagado_inicial')
                        if monto_pagado_inicial is None and pagado_por_usuario:
                            monto_pagado_inicial = nuevo_gasto.monto_total
                        for miembro in grupo.miembros.all():
                            monto_debido_miembro = nuevo_gasto.monto_por_persona
                            monto_pagado_miembro = Decimal('0')
                            if pagado_por_usuario and miembro == pagado_por_usuario:
                                monto_por_persona = nuevo_gasto.monto_por_persona
                                if monto_pagado_inicial:
                                    monto_pagado_miembro = min(monto_por_persona, monto_pagado_inicial)
                                else:
                                    monto_pagado_miembro = monto_por_persona
                            PagoGastoCompartido.objects.create(
                                gasto_compartido=nuevo_gasto,
                                miembro=miembro,
                                monto_debido=monto_debido_miembro,
                                monto_pagado=monto_pagado_miembro
                            )
                        # Notificaciones solo para la primera cuota
                        if i == cuota_actual:
                            url_gasto = reverse('gasto_compartido_detalle', args=[nuevo_gasto.pk])
                            for miembro in grupo.miembros.all():
                                if miembro != request.user:
                                    Notificacion.objects.create(
                                        usuario=miembro,
                                        mensaje=f"{request.user.username} ha añadido un nuevo gasto en '{grupo.nombre}': {nuevo_gasto.titulo}.",
                                        url_destino=url_gasto
                                    )
                    messages.success(request, f'Gasto compartido en cuotas creado exitosamente ({numero_cuotas} cuotas).')
                    return redirect('gastos_compartidos_lista')
                else:
                    gasto.save()
                    pagado_por_usuario = gasto.pagado_por
                    monto_pagado_inicial = form.cleaned_data.get('monto_pagado_inicial')
                    if monto_pagado_inicial is None and pagado_por_usuario:
                        monto_pagado_inicial = gasto.monto_total
                    for miembro in grupo.miembros.all():
                        monto_debido_miembro = gasto.monto_por_persona
                        monto_pagado_miembro = Decimal('0')
                        if pagado_por_usuario and miembro == pagado_por_usuario:
                            monto_por_persona = gasto.monto_por_persona
                            if monto_pagado_inicial:
                                monto_pagado_miembro = min(monto_por_persona, monto_pagado_inicial)
                            else:
                                monto_pagado_miembro = monto_por_persona
                        PagoGastoCompartido.objects.create(
                            gasto_compartido=gasto,
                            miembro=miembro,
                            monto_debido=monto_debido_miembro,
                            monto_pagado=monto_pagado_miembro
                        )
                    url_gasto = reverse('gasto_compartido_detalle', args=[gasto.pk])
                    for miembro in grupo.miembros.all():
                        if miembro != request.user:
                            Notificacion.objects.create(
                                usuario=miembro,
                                mensaje=f"{request.user.username} ha añadido un nuevo gasto en '{grupo.nombre}': {gasto.titulo}.",
                                url_destino=url_gasto
                            )
                    messages.success(request, 'Gasto compartido creado exitosamente.')
                    return redirect('gastos_compartidos_lista')
        else:
            grupo_id = request.GET.get('grupo')
            grupo_pre_seleccionado = None
            if grupo_id:
                try:
                    grupo_pre_seleccionado = get_object_or_404(GrupoGastosCompartidos, pk=grupo_id, miembros=request.user, activo=True)
                except:
                    messages.warning(request, 'El grupo especificado no existe o no tienes permisos. Selecciona otro grupo.')
            form = GastoCompartidoForm(user=request.user, grupo=grupo_pre_seleccionado)
        context = {
            'form': form,
            'grupos': grupos,
            'titulo': 'Nuevo Gasto Compartido',
        }
        return render(request, 'gastos_compartidos/form.html', context)
    except Exception as e:
        messages.error(request, f'Error al cargar la página: {str(e)}')
        return redirect('dashboard_gastos_compartidos')

@login_required
def gasto_compartido_editar(request, pk):
    """Editar un gasto compartido"""
    gasto = get_object_or_404(GastoCompartido, pk=pk, grupo__miembros=request.user, activo=True)
    
    if request.method == 'POST':
        form = GastoCompartidoForm(request.POST, request.FILES, instance=gasto, user=request.user, grupo=gasto.grupo)
        if form.is_valid():
            form.save()
            messages.success(request, 'Gasto compartido actualizado exitosamente.')
            return redirect('gastos_compartidos_lista')
    else:
        form = GastoCompartidoForm(instance=gasto, user=request.user, grupo=gasto.grupo)
    
    context = {
        'form': form,
        'gasto': gasto,
        'titulo': 'Editar Gasto Compartido',
    }
    return render(request, 'gastos_compartidos/form.html', context)

@login_required
def gasto_compartido_eliminar(request, pk):
    """Eliminar un gasto compartido - Versión simple sin AJAX"""
    try:
        gasto = get_object_or_404(GastoCompartido, pk=pk, activo=True)
        
        # Verificar permisos
        if request.user not in gasto.grupo.miembros.all():
            messages.error(request, 'No tienes permisos para eliminar este gasto.')
            return redirect('gastos_compartidos_lista')
        
        # Si es POST, eliminar directamente
        if request.method == 'POST':
            # Marcar el gasto como inactivo (eliminación lógica)
            gasto.activo = False
            gasto.save()
            messages.success(request, 'Gasto eliminado exitosamente.')
            return redirect('gastos_compartidos_lista')
        
        # Si es GET, mostrar confirmación
        context = {
            'gasto': gasto,
            'pagos_count': PagoGastoCompartido.objects.filter(gasto_compartido=gasto).count(),
            'pagados_count': PagoGastoCompartido.objects.filter(gasto_compartido=gasto, estado='pagado').count(),
        }
        return render(request, 'gastos_compartidos/eliminar.html', context)
        
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('gastos_compartidos_lista')

@login_required
def gasto_compartido_detalle(request, pk):
    """Ver detalles de un gasto compartido"""
    gasto = get_object_or_404(GastoCompartido, pk=pk, grupo__miembros=request.user, activo=True)
    pagos = PagoGastoCompartido.objects.filter(gasto_compartido=gasto).select_related('miembro')
    
    # Calcular estadísticas de pagos
    total_pagado = sum(pago.monto_pagado for pago in pagos)
    total_debido = sum(pago.monto_debido for pago in pagos)
    total_pendiente = total_debido - total_pagado
    
    # Calcular porcentaje de pago
    if gasto.monto_total > 0:
        porcentaje_pagado = (total_pagado / gasto.monto_total) * 100
    else:
        porcentaje_pagado = 0
    
    # Contar pagos por estado
    pagos_pendientes = pagos.filter(estado='pendiente').count()
    pagos_pagados = pagos.filter(estado='pagado').count()
    pagos_vencidos = pagos.filter(estado='vencido').count()
    
    # Determinar el estado general del gasto
    if total_pendiente == 0:
        estado_general = 'pagado'
    elif gasto.fecha_vencimiento and gasto.fecha_vencimiento < date.today():
        estado_general = 'vencido'
    else:
        estado_general = 'pendiente'
    
    context = {
        'gasto': gasto,
        'pagos': pagos,
        'total_pagado': total_pagado,
        'total_debido': total_debido,
        'total_pendiente': total_pendiente,
        'porcentaje_pagado': porcentaje_pagado,
        'pagos_pendientes': pagos_pendientes,
        'pagos_pagados': pagos_pagados,
        'pagos_vencidos': pagos_vencidos,
        'estado_general': estado_general,
        'total_miembros': pagos.count(),
    }
    return render(request, 'gastos_compartidos/detalle.html', context)

@login_required
def pago_gasto_compartido_editar(request, pk):
    """Editar el pago de un miembro para un gasto compartido"""
    pago = get_object_or_404(PagoGastoCompartido, pk=pk, miembro=request.user)
    
    if request.method == 'POST':
        form = PagoGastoCompartidoForm(request.POST, instance=pago)
        if form.is_valid():
            # El save() del modelo ya actualiza automáticamente el estado
            form.save()
            messages.success(request, 'Pago actualizado exitosamente.')
            return redirect('gasto_compartido_detalle', pk=pago.gasto_compartido.pk)
    else:
        form = PagoGastoCompartidoForm(instance=pago)
    
    context = {
        'form': form,
        'pago': pago,
        'titulo': 'Editar Pago',
    }
    return render(request, 'gastos_compartidos/pago_form.html', context)

# Vista para histórico de grupos y gastos eliminados
@login_required
def historico_gastos_compartidos(request):
    """Vista para ver el histórico de grupos y gastos eliminados"""
    
    # Si el usuario es staff, mostrar todo el histórico
    if request.user.is_staff or request.user.is_superuser:
        grupos_inactivos = GrupoGastosCompartidos.objects.filter(
            activo=False
        ).order_by('-fecha_modificacion')
        
        gastos_inactivos = GastoCompartido.objects.filter(
            activo=False
        ).select_related('grupo', 'pagado_por').order_by('-fecha_modificacion')
        
        es_staff = True
    else:
        # Usuario normal: solo ver su propio histórico
        grupos_inactivos = GrupoGastosCompartidos.objects.filter(
            Q(miembros=request.user) | Q(creador=request.user),
            activo=False
        ).order_by('-fecha_modificacion')
        
        gastos_inactivos = GastoCompartido.objects.filter(
            grupo__miembros=request.user,
            activo=False
        ).select_related('grupo', 'pagado_por').order_by('-fecha_modificacion')
        
        es_staff = False
    
    context = {
        'grupos_inactivos': grupos_inactivos,
        'gastos_inactivos': gastos_inactivos,
        'total_grupos_inactivos': grupos_inactivos.count(),
        'total_gastos_inactivos': gastos_inactivos.count(),
        'es_staff': es_staff,
    }
    return render(request, 'gastos_compartidos/historico.html', context)

@login_required
def saldos_grupo(request, pk):
    """Mostrar saldos de un grupo específico"""
    grupo = get_object_or_404(GrupoGastosCompartidos, pk=pk, miembros=request.user, activo=True)
    
    # Obtener todos los gastos del grupo
    gastos = GastoCompartido.objects.filter(grupo=grupo, activo=True).select_related('pagado_por')
    
    # Calcular saldos por miembro
    saldos_miembros = []
    for miembro in grupo.miembros.all():
        # Obtener pagos del miembro
        pagos_miembro = PagoGastoCompartido.objects.filter(
            miembro=miembro,
            gasto_compartido__grupo=grupo,
            gasto_compartido__activo=True
        )
        
        total_debido = sum(pago.monto_debido for pago in pagos_miembro)
        total_pagado = sum(pago.monto_pagado for pago in pagos_miembro)
        saldo = total_debido - total_pagado
        
        saldos_miembros.append({
            'miembro': miembro,
            'total_debido': total_debido,
            'total_pagado': total_pagado,
            'saldo': saldo,
            'estado': 'al día' if saldo == 0 else 'debe' if saldo > 0 else 'a favor'
        })
    
    # Ordenar por saldo (deudores primero)
    saldos_miembros.sort(key=lambda x: x['saldo'], reverse=True)
    
    context = {
        'grupo': grupo,
        'saldos_miembros': saldos_miembros,
        'total_gastos': gastos.count(),
        'total_monto': sum(gasto.monto_total for gasto in gastos),
    }
    
    return render(request, 'gastos_compartidos/saldos_grupo.html', context)

@login_required
def miembros_grupo(request, pk):
    """Gestionar miembros de un grupo"""
    grupo = get_object_or_404(GrupoGastosCompartidos, pk=pk, miembros=request.user, activo=True)
    
    if request.method == 'POST':
        # Agregar nuevo miembro
        username = request.POST.get('username')
        if username:
            try:
                nuevo_miembro = User.objects.get(username=username, is_active=True)
                if nuevo_miembro not in grupo.miembros.all():
                    grupo.miembros.add(nuevo_miembro)
                    messages.success(request, f'{nuevo_miembro.username} ha sido agregado al grupo.')
                else:
                    messages.warning(request, f'{nuevo_miembro.username} ya es miembro del grupo.')
            except User.DoesNotExist:
                messages.error(request, f'El usuario {username} no existe.')
        
        # Remover miembro
        miembro_id = request.POST.get('remover_miembro')
        if miembro_id:
            try:
                miembro_a_remover = User.objects.get(id=miembro_id)
                if miembro_a_remover in grupo.miembros.all() and miembro_a_remover != request.user:
                    grupo.miembros.remove(miembro_a_remover)
                    messages.success(request, f'{miembro_a_remover.username} ha sido removido del grupo.')
                else:
                    messages.error(request, 'No puedes removerte a ti mismo del grupo.')
            except User.DoesNotExist:
                messages.error(request, 'Usuario no encontrado.')
    
    # Obtener lista de usuarios disponibles para agregar
    usuarios_disponibles = User.objects.filter(is_active=True).exclude(
        id__in=grupo.miembros.values_list('id', flat=True)
    )
    
    context = {
        'grupo': grupo,
        'miembros': grupo.miembros.all(),
        'usuarios_disponibles': usuarios_disponibles,
    }
    
    return render(request, 'gastos_compartidos/miembros_grupo.html', context)

# APIs para Gastos Compartidos
@login_required
def api_grupo_miembros(request, grupo_id):
    """API para obtener miembros de un grupo"""
    try:
        grupo = get_object_or_404(GrupoGastosCompartidos, pk=grupo_id, miembros=request.user, activo=True)
        miembros = grupo.miembros.values('id', 'username', 'first_name', 'last_name')
        return JsonResponse({'miembros': list(miembros)})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def api_crear_grupo(request, grupo_id):
    """API para crear un grupo"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            nombre = data.get('nombre')
            descripcion = data.get('descripcion', '')
            
            if not nombre:
                return JsonResponse({'error': 'El nombre es requerido'}, status=400)
            
            grupo = GrupoGastosCompartidos.objects.create(
                nombre=nombre,
                descripcion=descripcion,
                creado_por=request.user
            )
            grupo.miembros.add(request.user)
            
            return JsonResponse({
                'id': grupo.id,
                'nombre': grupo.nombre,
                'descripcion': grupo.descripcion
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required
def api_crear_gasto(request):
    """API para crear un gasto compartido"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            grupo_id = data.get('grupo_id')
            titulo = data.get('titulo')
            monto_total = data.get('monto_total')
            
            if not all([grupo_id, titulo, monto_total]):
                return JsonResponse({'error': 'Todos los campos son requeridos'}, status=400)
            
            grupo = get_object_or_404(GrupoGastosCompartidos, pk=grupo_id, miembros=request.user, activo=True)
            
            gasto = GastoCompartido.objects.create(
                grupo=grupo,
                titulo=titulo,
                monto_total=monto_total,
                pagado_por=request.user
            )
            
            # Crear pagos para cada miembro
            for miembro in grupo.miembros.all():
                monto_debido = gasto.monto_por_persona
                monto_pagado = Decimal('0')
                
                if miembro == request.user:
                    monto_pagado = monto_debido
                
                PagoGastoCompartido.objects.create(
                    gasto_compartido=gasto,
                    miembro=miembro,
                    monto_debido=monto_debido,
                    monto_pagado=monto_pagado
                )
            
            return JsonResponse({
                'id': gasto.id,
                'titulo': gasto.titulo,
                'monto_total': str(gasto.monto_total)
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

@login_required
def api_editar_pago(request):
    """API para editar un pago"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            pago_id = data.get('pago_id')
            monto_pagado = data.get('monto_pagado')
            
            if not all([pago_id, monto_pagado]):
                return JsonResponse({'error': 'Todos los campos son requeridos'}, status=400)
            
            pago = get_object_or_404(PagoGastoCompartido, pk=pago_id, miembro=request.user)
            pago.monto_pagado = Decimal(monto_pagado)
            pago.save()
            
            return JsonResponse({
                'id': pago.id,
                'monto_pagado': str(pago.monto_pagado),
                'estado': pago.estado
            })
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Método no permitido'}, status=405)

# Vistas para Notificaciones
@login_required
def lista_notificaciones(request):
    """Lista de notificaciones del usuario"""
    notificaciones = Notificacion.objects.filter(
        usuario=request.user
    ).order_by('-fecha_creacion')
    
    # Si es una petición AJAX, devolver JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        notificaciones_data = []
        for notif in notificaciones[:10]:  # Solo las últimas 10
            notificaciones_data.append({
                'id': notif.id,
                'mensaje': notif.mensaje,
                'url_destino': notif.url_destino or '',
                'leida': notif.leida,
                'fecha_creacion': notif.fecha_creacion.isoformat()
            })
        
        return JsonResponse({
            'notificaciones': notificaciones_data,
            'total': notificaciones.count(),
            'no_leidas': notificaciones.filter(leida=False).count()
        })
    
    # Paginación para vista normal
    paginator = Paginator(notificaciones, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_notificaciones': notificaciones.count(),
        'notificaciones_no_leidas': notificaciones.filter(leida=False).count(),
    }
    
    return render(request, 'notificaciones/lista.html', context)

@login_required
def marcar_notificacion_leida(request, pk):
    """Marcar una notificación como leída"""
    notificacion = get_object_or_404(Notificacion, pk=pk, usuario=request.user)
    notificacion.leida = True
    notificacion.save()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    return redirect('lista_notificaciones')

@login_required
def marcar_todas_notificaciones_leidas(request):
    """Marcar todas las notificaciones como leídas"""
    Notificacion.objects.filter(usuario=request.user, leida=False).update(leida=True)
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    messages.success(request, 'Todas las notificaciones han sido marcadas como leídas.')
    return redirect('lista_notificaciones')

@login_required
def notificaciones_count(request):
    """Obtener el conteo de notificaciones no leídas"""
    count = Notificacion.objects.filter(usuario=request.user, leida=False).count()
    return JsonResponse({'count': count})

@login_required
def eliminar_notificacion(request, pk):
    """Eliminar una notificación"""
    notificacion = get_object_or_404(Notificacion, pk=pk, usuario=request.user)
    notificacion.delete()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    messages.success(request, 'Notificación eliminada correctamente.')
    return redirect('lista_notificaciones')

# APIs adicionales
@login_required
def api_cuentas_usuario(request, user_id):
    """API para obtener cuentas de un usuario"""
    try:
        if request.user.is_staff or request.user.id == user_id:
            cuentas = Cuenta.objects.filter(usuario_id=user_id, activa=True).values('id', 'nombre', 'saldo_actual')
            return JsonResponse({'cuentas': list(cuentas)})
        else:
            return JsonResponse({'error': 'No tienes permisos'}, status=403)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

@login_required
def api_grupo_info(request, grupo_id):
    """API para obtener información de un grupo"""
    try:
        grupo = get_object_or_404(GrupoGastosCompartidos, pk=grupo_id, miembros=request.user, activo=True)
        
        # Calcular estadísticas del grupo
        gastos = GastoCompartido.objects.filter(grupo=grupo, activo=True)
        total_gastos = gastos.count()
        total_monto = sum(gasto.monto_total for gasto in gastos)
        
        # Saldo del usuario en el grupo
        pagos_usuario = PagoGastoCompartido.objects.filter(
            miembro=request.user,
            gasto_compartido__grupo=grupo,
            gasto_compartido__activo=True
        )
        saldo_usuario = sum(pago.monto_debido - pago.monto_pagado for pago in pagos_usuario)
        
        return JsonResponse({
            'id': grupo.id,
            'nombre': grupo.nombre,
            'descripcion': grupo.descripcion,
            'total_gastos': total_gastos,
            'total_monto': str(total_monto),
            'saldo_usuario': str(saldo_usuario),
            'miembros_count': grupo.miembros.count()
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

# Vistas para Reportes de Gastos Compartidos
@login_required
def reportes_gastos_compartidos(request):
    """Vista principal para generar reportes de gastos compartidos"""
    # Obtener grupos del usuario
    grupos = GrupoGastosCompartidos.objects.filter(miembros=request.user, activo=True).order_by('nombre')
    
    # Obtener meses disponibles para reportes
    meses_disponibles = []
    for grupo in grupos:
        gastos_grupo = GastoCompartido.objects.filter(grupo=grupo, activo=True).dates('fecha', 'month')
        for fecha in gastos_grupo:
            mes_info = {
                'año': fecha.year,
                'mes': fecha.month,
                'nombre': fecha.strftime('%B %Y'),
                'grupo_id': grupo.id,
                'grupo_nombre': grupo.nombre
            }
            if mes_info not in meses_disponibles:
                meses_disponibles.append(mes_info)
    
    # Ordenar por fecha (más reciente primero)
    meses_disponibles.sort(key=lambda x: (x['año'], x['mes']), reverse=True)
    
    context = {
        'grupos': grupos,
        'meses_disponibles': meses_disponibles,
        'titulo': 'Reportes de Gastos Compartidos'
    }
    return render(request, 'reportes/reportes_gastos_compartidos.html', context)

@login_required
def generar_reporte_gastos_compartidos(request):
    """Generar reporte de gastos compartidos en formato PDF o Excel"""
    if request.method != 'POST':
        return redirect('reportes_gastos_compartidos')
    
    # Obtener parámetros del formulario
    grupo_id = request.POST.get('grupo')
    fecha_desde = request.POST.get('fecha_desde')
    fecha_hasta = request.POST.get('fecha_hasta')
    formato = request.POST.get('formato', 'pdf')
    
    # Validar parámetros
    if not grupo_id or not fecha_desde or not fecha_hasta:
        messages.error(request, 'Todos los campos son requeridos.')
        return redirect('reportes_gastos_compartidos')
    
    try:
        grupo = get_object_or_404(GrupoGastosCompartidos, pk=grupo_id, miembros=request.user, activo=True)
        fecha_desde = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        fecha_hasta = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
    except (ValueError, GrupoGastosCompartidos.DoesNotExist):
        messages.error(request, 'Parámetros inválidos.')
        return redirect('reportes_gastos_compartidos')
    
    # Obtener datos del reporte
    gastos = GastoCompartido.objects.filter(
        grupo=grupo,
        activo=True,
        fecha__gte=fecha_desde,
        fecha__lte=fecha_hasta
    ).select_related('pagado_por', 'cuenta_pago').order_by('fecha')
    
    # Obtener pagos de todos los miembros
    pagos = PagoGastoCompartido.objects.filter(
        gasto_compartido__in=gastos
    ).select_related('miembro', 'gasto_compartido')
    
    # Calcular estadísticas
    total_gastos = gastos.count()
    total_monto = sum(gasto.monto_total for gasto in gastos)
    
    # Calcular saldos por miembro
    saldos_miembros = {}
    for miembro in grupo.miembros.all():
        pagos_miembro = pagos.filter(miembro=miembro)
        total_debido = sum(pago.monto_debido for pago in pagos_miembro)
        total_pagado = sum(pago.monto_pagado for pago in pagos_miembro)
        saldo = total_debido - total_pagado
        
        saldos_miembros[miembro.username] = {
            'miembro': miembro,
            'total_debido': total_debido,
            'total_pagado': total_pagado,
            'saldo': saldo,
            'estado': 'al día' if saldo == 0 else 'debe' if saldo > 0 else 'a favor'
        }
    
    # Preparar datos para el reporte
    datos_reporte = {
        'grupo': grupo,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
        'gastos': gastos,
        'pagos': pagos,
        'saldos_miembros': saldos_miembros,
        'total_gastos': total_gastos,
        'total_monto': total_monto,
        'miembros_count': grupo.miembros.count(),
        'generado_por': request.user,
        'fecha_generacion': date.today()
    }
    
    # Generar reporte según formato
    if formato == 'pdf':
        return generar_pdf_gastos_compartidos(datos_reporte)
    elif formato == 'excel':
        return generar_excel_gastos_compartidos(datos_reporte)
    else:
        messages.error(request, 'Formato no válido.')
        return redirect('reportes_gastos_compartidos')

def generar_pdf_gastos_compartidos(datos):
    """Generar reporte PDF de gastos compartidos (versión con saldos netos entre usuarios)"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER
    )
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20
    )
    normal_style = styles['Normal']

    # Título y periodo
    story.append(Paragraph(f"Reporte de Gastos Compartidos", title_style))
    story.append(Paragraph(f"Grupo: {datos['grupo'].nombre}", subtitle_style))
    story.append(Paragraph(f"Período: {datos['fecha_desde'].strftime('%d/%m/%Y')} - {datos['fecha_hasta'].strftime('%d/%m/%Y')}", normal_style))
    story.append(Spacer(1, 20))

    # Detalle de gastos
    story.append(Paragraph("Detalle de Gastos", subtitle_style))
    if datos['gastos']:
        gastos_data = [['Fecha', 'Título', 'Monto', 'Pagado por', 'Cuenta', 'Monto por Persona']]
        for gasto in datos['gastos']:
            gastos_data.append([
                gasto.fecha.strftime('%d/%m/%Y'),
                gasto.titulo[:30] + '...' if len(gasto.titulo) > 30 else gasto.titulo,
                f"${gasto.monto_total:,.2f}",
                gasto.pagado_por.get_full_name() or gasto.pagado_por.username if gasto.pagado_por else 'N/A',
                gasto.cuenta_pago.nombre if gasto.cuenta_pago else 'N/A',
                f"${gasto.monto_por_persona:,.2f}"
            ])
        gastos_table = Table(gastos_data, colWidths=[0.8*inch, 1.8*inch, 0.8*inch, 1.2*inch, 1*inch, 1*inch])
        gastos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ]))
        story.append(gastos_table)
        story.append(Spacer(1, 20))
        # Calcular saldos netos entre usuarios
        miembros = list(datos['grupo'].miembros.all())
        pagos = datos['pagos']
        # Diccionario: {usuario: {otro_usuario: saldo}}
        saldos = {m: {o: 0 for o in miembros if o != m} for m in miembros}
        # Para cada gasto, calcular cuánto debe cada uno al pagador
        for gasto in datos['gastos']:
            pagador = gasto.pagado_por
            if not pagador:
                continue
            for miembro in miembros:
                if miembro != pagador:
                    monto = float(gasto.monto_por_persona)
                    saldos[miembro][pagador] += monto
        # Netear saldos (solo mostrar deudores netos)
        saldos_neto = []
        for a in miembros:
            for b in miembros:
                if a != b:
                    neto = saldos[a][b] - saldos[b][a]
                    if neto > 0:
                        saldos_neto.append((a, b, neto))
        if saldos_neto:
            story.append(Paragraph("Saldos entre usuarios (quién le debe a quién):", subtitle_style))
            tabla_saldos = [["Deudor", "Acreedor", "Monto"]]
            for deudor, acreedor, monto in saldos_neto:
                tabla_saldos.append([
                    deudor.get_full_name() or deudor.username,
                    acreedor.get_full_name() or acreedor.username,
                    f"${monto:,.2f}"
                ])
            tabla = Table(tabla_saldos, colWidths=[2*inch, 2*inch, 1.2*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            story.append(tabla)
    else:
        story.append(Paragraph("No hay gastos en el período seleccionado.", normal_style))
    story.append(Spacer(1, 20))
    story.append(Paragraph(f"Reporte generado por: {datos['generado_por'].get_full_name() or datos['generado_por'].username}", normal_style))
    story.append(Paragraph(f"Fecha de generación: {datos['fecha_generacion'].strftime('%d/%m/%Y')}", normal_style))
    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_gastos_compartidos_{datos["grupo"].nombre}_{datos["fecha_desde"].strftime("%Y%m%d")}_{datos["fecha_hasta"].strftime("%Y%m%d")}.pdf"'
    return response

def generar_excel_gastos_compartidos(datos):
    """Generar reporte Excel de gastos compartidos (versión con saldos netos entre usuarios)"""
    buffer = io.BytesIO()
    workbook = xlsxwriter.Workbook(buffer)
    header_format = workbook.add_format({
        'bold': True,
        'text_wrap': True,
        'valign': 'top',
        'fg_color': '#D7E4BC',
        'border': 1
    })
    cell_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter'
    })
    money_format = workbook.add_format({
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'num_format': '$#,##0.00'
    })
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'align': 'center',
        'valign': 'vcenter'
    })
    worksheet = workbook.add_worksheet('Detalle de Gastos')
    worksheet.merge_range('A1:F1', f'Reporte de Gastos Compartidos - {datos["grupo"].nombre}', title_format)
    worksheet.merge_range('A2:F2', f'Período: {datos["fecha_desde"].strftime("%d/%m/%Y")} - {datos["fecha_hasta"].strftime("%d/%m/%Y")}', title_format)
    headers = ['Fecha', 'Título', 'Monto Total', 'Pagado por', 'Cuenta', 'Monto por Persona']
    for col, header in enumerate(headers):
        worksheet.write(3, col, header, header_format)
    row = 4
    pagos_por_usuario = {}
    miembros = list(datos['grupo'].miembros.all())
    for gasto in datos['gastos']:
        worksheet.write(row, 0, gasto.fecha.strftime('%d/%m/%Y'), cell_format)
        worksheet.write(row, 1, gasto.titulo, cell_format)
        worksheet.write(row, 2, gasto.monto_total, money_format)
        pagador = gasto.pagado_por.get_full_name() or gasto.pagado_por.username if gasto.pagado_por else 'N/A'
        worksheet.write(row, 3, pagador, cell_format)
        worksheet.write(row, 4, gasto.cuenta_pago.nombre if gasto.cuenta_pago else 'N/A', cell_format)
        worksheet.write(row, 5, gasto.monto_por_persona, money_format)
        if gasto.pagado_por:
            pagos_por_usuario[pagador] = pagos_por_usuario.get(pagador, 0) + float(gasto.monto_total)
        row += 1
    # Calcular saldos netos entre usuarios
    saldos = {m: {o: 0 for o in miembros if o != m} for m in miembros}
    for gasto in datos['gastos']:
        pagador = gasto.pagado_por
        if not pagador:
            continue
        for miembro in miembros:
            if miembro != pagador:
                monto = float(gasto.monto_por_persona)
                saldos[miembro][pagador] += monto
    saldos_neto = []
    for a in miembros:
        for b in miembros:
            if a != b:
                neto = saldos[a][b] - saldos[b][a]
                if neto > 0:
                    saldos_neto.append((a, b, neto))
    if saldos_neto:
        worksheet.write(row+1, 1, 'Deudor', header_format)
        worksheet.write(row+1, 2, 'Acreedor', header_format)
        worksheet.write(row+1, 3, 'Monto', header_format)
        for i, (deudor, acreedor, monto) in enumerate(saldos_neto):
            worksheet.write(row+2+i, 1, deudor.get_full_name() or deudor.username, cell_format)
            worksheet.write(row+2+i, 2, acreedor.get_full_name() or acreedor.username, cell_format)
            worksheet.write(row+2+i, 3, monto, money_format)
    # Gráfico de torta
    if pagos_por_usuario:
        chart = workbook.add_chart({'type': 'pie'})
        usuarios = list(pagos_por_usuario.keys())
        valores = list(pagos_por_usuario.values())
        for i, usuario_nombre in enumerate(usuarios):
            worksheet.write(row+20+i, 1, usuario_nombre)
            worksheet.write(row+20+i, 2, valores[i])
        chart.add_series({
            'name': 'Proporción de gastos pagados',
            'categories': ['Detalle de Gastos', row+20, 1, row+20+len(usuarios)-1, 1],
            'values':     ['Detalle de Gastos', row+20, 2, row+20+len(usuarios)-1, 2],
            'data_labels': {'percentage': True}
        })
        chart.set_title({'name': 'Gastos pagados por usuario'})
        worksheet.insert_chart(row+20, 4, chart, {'x_offset': 25, 'y_offset': 10})
    worksheet.set_column('A:A', 12)
    worksheet.set_column('B:B', 25)
    worksheet.set_column('C:C', 15)
    worksheet.set_column('D:D', 20)
    worksheet.set_column('E:E', 15)
    worksheet.set_column('F:F', 15)
    worksheet.write(row+40, 0, f'Reporte generado por: {datos["generado_por"].get_full_name() or datos["generado_por"].username}')
    #para que ande en server borar si anda malworksheet.write(row+40, 0, f'Reporte generado por: {datos['generado_por'].get_full_name() or datos['generado_por'].username}')
    worksheet.write(row+41, 0, f'Fecha de generación: {datos["fecha_generacion"].strftime("%d/%m/%Y")}')
    workbook.close()
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="reporte_gastos_compartidos_{datos["grupo"].nombre}_{datos["fecha_desde"].strftime("%Y%m%d")}_{datos["fecha_hasta"].strftime("%Y%m%d")}.xlsx"'
    return response

@login_required
def reportes_corte_mes(request):
    """Vista para mostrar el formulario de reportes de corte de mes"""
    # Obtener cortes disponibles del usuario
    cortes_disponibles = CorteMes.objects.filter(usuario=request.user).order_by('-fecha_corte')
    
    context = {
        'cortes_disponibles': cortes_disponibles,
    }
    return render(request, 'reportes/reportes_corte_mes.html', context)

@login_required
def generar_reporte_corte_mes(request):
    """Vista para generar reportes de corte de mes"""
    if request.method != 'POST':
        return redirect('reportes_corte_mes')
    
    # Obtener parámetros del formulario
    corte_id = request.POST.get('corte_id')
    formato = request.POST.get('formato', 'pdf')
    
    try:
        corte = CorteMes.objects.get(id=corte_id, usuario=request.user)
    except CorteMes.DoesNotExist:
        messages.error(request, 'Corte de mes no encontrado.')
        return redirect('reportes_corte_mes')
    
    # Preparar datos para el reporte
    datos = {
        'corte': corte,
        'usuario': request.user,
        'fecha_generacion': timezone.now(),
    }
    
    # Generar reporte según el formato
    if formato == 'pdf':
        response = generar_pdf_corte_mes(datos)
        filename = f"corte_mes_{corte.mes_cortado}_{corte.año_cortado}.pdf"
    else:  # excel
        response = generar_excel_corte_mes(datos)
        filename = f"corte_mes_{corte.mes_cortado}_{corte.año_cortado}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def generar_pdf_corte_mes(datos):
    """Genera un reporte PDF del corte de mes"""
    corte = datos['corte']
    usuario = datos['usuario']
    
    # Crear el buffer para el PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Centrado
        textColor=colors.darkblue
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=20,
        textColor=colors.darkblue
    )
    
    normal_style = styles['Normal']
    
    # Título del reporte
    elements.append(Paragraph(f"Reporte de Corte de Mes - {corte.mes_nombre} {corte.año_cortado}", title_style))
    elements.append(Paragraph(f"Usuario: {usuario.get_full_name() or usuario.username}", normal_style))
    elements.append(Paragraph(f"Fecha de generación: {datos['fecha_generacion'].strftime('%d/%m/%Y %H:%M')}", normal_style))
    elements.append(Spacer(1, 20))
    
    # Resumen ejecutivo
    elements.append(Paragraph("Resumen Ejecutivo", subtitle_style))
    
    # Tabla de resumen
    resumen_data = [
        ['Concepto', 'Monto'],
        ['Total Ingresos', f"${corte.total_ingresos:,.2f}"],
        ['Total Gastos', f"${corte.total_gastos:,.2f}"],
        ['Balance del Mes', f"${corte.balance_mes:,.2f}"],
    ]
    
    resumen_table = Table(resumen_data, colWidths=[200, 100])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),  # Alinear montos a la derecha
    ]))
    elements.append(resumen_table)
    elements.append(Spacer(1, 20))
    
    # Saldos de cuentas
    elements.append(Paragraph("Saldos de Cuentas al Corte", subtitle_style))
    
    saldos_data = [['Cuenta', 'Saldo']]
    
    # Verificar el formato de los datos
    if isinstance(corte.saldos_cuentas, dict):
        # Si es el formato nuevo (JSON con datos completos)
        if 'saldos_cuentas' in corte.saldos_cuentas:
            # Formato nuevo: {'saldos_cuentas': {cuenta_id: saldo}, 'gastos_compartidos': {...}}
            for cuenta_id, saldo in corte.saldos_cuentas['saldos_cuentas'].items():
                try:
                    # Asegurar que cuenta_id sea un número
                    cuenta_id_int = int(cuenta_id)
                    cuenta = Cuenta.objects.get(id=cuenta_id_int)
                    saldos_data.append([cuenta.nombre, f"${saldo:,.2f}"])
                except (ValueError, Cuenta.DoesNotExist):
                    continue
        else:
            # Formato antiguo: {cuenta_id: saldo}
            for cuenta_id, saldo in corte.saldos_cuentas.items():
                try:
                    # Asegurar que cuenta_id sea un número
                    cuenta_id_int = int(cuenta_id)
                    cuenta = Cuenta.objects.get(id=cuenta_id_int)
                    saldos_data.append([cuenta.nombre, f"${saldo:,.2f}"])
                except (ValueError, Cuenta.DoesNotExist):
                    continue
    
    if len(saldos_data) > 1:
        saldos_table = Table(saldos_data, colWidths=[200, 100])
        saldos_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ]))
        elements.append(saldos_table)
    else:
        elements.append(Paragraph("No hay datos de saldos de cuentas disponibles.", normal_style))
    
    elements.append(Spacer(1, 20))
    
    # Gastos compartidos (si están disponibles)
    if isinstance(corte.saldos_cuentas, dict) and 'gastos_compartidos' in corte.saldos_cuentas:
        gastos_compartidos = corte.saldos_cuentas['gastos_compartidos']
        
        elements.append(Paragraph("Gastos Compartidos del Mes", subtitle_style))
        
        # Resumen de gastos compartidos
        gc_resumen_data = [
            ['Concepto', 'Monto'],
            ['Total Gastos Compartidos', f"${gastos_compartidos.get('total_gastos_compartidos_mes', 0):,.2f}"],
            ['Gastos Pagados por Usuario', f"${gastos_compartidos.get('gastos_pagados_por_usuario', 0):,.2f}"],
            ['Pagos Realizados por Usuario', f"${gastos_compartidos.get('pagos_usuario_mes', 0):,.2f}"],
            ['Total Saldo Pendiente', f"${gastos_compartidos.get('total_saldo_pendiente', 0):,.2f}"],
        ]
        
        gc_resumen_table = Table(gc_resumen_data, colWidths=[200, 100])
        gc_resumen_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
        ]))
        elements.append(gc_resumen_table)
        elements.append(Spacer(1, 20))
        
        # Detalle de saldos por grupo
        if 'saldos_pendientes_grupos' in gastos_compartidos:
            elements.append(Paragraph("Saldos Pendientes por Grupo", subtitle_style))
            
            gc_saldos_data = [
                ['Grupo', 'Total Debido', 'Total Pagado', 'Saldo Pendiente', 'Miembros']
            ]
            
            for grupo_id, datos_grupo in gastos_compartidos['saldos_pendientes_grupos'].items():
                gc_saldos_data.append([
                    datos_grupo['nombre'],
                    f"${datos_grupo['total_debido']:,.2f}",
                    f"${datos_grupo['total_pagado']:,.2f}",
                    f"${datos_grupo['saldo_pendiente']:,.2f}",
                    str(datos_grupo['cantidad_miembros'])
                ])
            
            gc_saldos_table = Table(gc_saldos_data, colWidths=[120, 80, 80, 80, 60])
            gc_saldos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
            ]))
            elements.append(gc_saldos_table)
    
    # Información adicional
    elements.append(Spacer(1, 20))
    elements.append(Paragraph("Información Adicional", subtitle_style))
    elements.append(Paragraph(f"Fecha de corte: {corte.fecha_corte.strftime('%d/%m/%Y')}", normal_style))
    elements.append(Paragraph(f"Mantener saldos: {'Sí' if corte.mantener_saldos else 'No'}", normal_style))
    elements.append(Paragraph(f"Fecha de creación del corte: {corte.fecha_creacion.strftime('%d/%m/%Y %H:%M')}", normal_style))
    
    # Construir el PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Crear la respuesta HTTP
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    return response

def generar_excel_corte_mes(datos):
    """Genera un reporte Excel del corte de mes"""
    corte = datos['corte']
    usuario = datos['usuario']
    
    # Crear el workbook
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    
    # Formatos
    title_format = workbook.add_format({
        'bold': True,
        'font_size': 16,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#2E86AB',
        'font_color': 'white',
        'border': 1
    })
    
    subtitle_format = workbook.add_format({
        'bold': True,
        'font_size': 12,
        'bg_color': '#A23B72',
        'font_color': 'white',
        'border': 1
    })
    
    header_format = workbook.add_format({
        'bold': True,
        'font_size': 11,
        'bg_color': '#F18F01',
        'font_color': 'white',
        'border': 1,
        'align': 'center'
    })
    
    money_format = workbook.add_format({
        'num_format': '$#,##0.00',
        'border': 1,
        'align': 'right'
    })
    
    normal_format = workbook.add_format({
        'border': 1,
        'align': 'left'
    })
    
    # Hoja 1: Resumen
    worksheet = workbook.add_worksheet('Resumen')
    
    # Título
    worksheet.merge_range('A1:D1', f'Reporte de Corte de Mes - {corte.mes_nombre} {corte.año_cortado}', title_format)
    worksheet.write('A3', 'Usuario:', normal_format)
    worksheet.write('B3', usuario.get_full_name() or usuario.username, normal_format)
    worksheet.write('A4', 'Fecha de generación:', normal_format)
    worksheet.write('B4', datos['fecha_generacion'].strftime('%d/%m/%Y %H:%M'), normal_format)
    
    # Resumen ejecutivo
    worksheet.merge_range('A6:D6', 'Resumen Ejecutivo', subtitle_format)
    worksheet.write('A7', 'Concepto', header_format)
    worksheet.write('B7', 'Monto', header_format)
    
    resumen_data = [
        ['Total Ingresos', float(corte.total_ingresos)],
        ['Total Gastos', float(corte.total_gastos)],
        ['Balance del Mes', float(corte.balance_mes)],
    ]
    
    for i, (concepto, monto) in enumerate(resumen_data):
        worksheet.write(f'A{8+i}', concepto, normal_format)
        worksheet.write(f'B{8+i}', monto, money_format)
    
    # Saldos de cuentas
    row = 12
    worksheet.merge_range(f'A{row}:D{row}', 'Saldos de Cuentas al Corte', subtitle_format)
    row += 1
    
    worksheet.write(f'A{row}', 'Cuenta', header_format)
    worksheet.write(f'B{row}', 'Saldo', header_format)
    row += 1
    
    # Verificar el formato de los datos
    if isinstance(corte.saldos_cuentas, dict):
        # Si es el formato nuevo (JSON con datos completos)
        if 'saldos_cuentas' in corte.saldos_cuentas:
            # Formato nuevo: {'saldos_cuentas': {cuenta_id: saldo}, 'gastos_compartidos': {...}}
            for cuenta_id, saldo in corte.saldos_cuentas['saldos_cuentas'].items():
                try:
                    # Asegurar que cuenta_id sea un número
                    cuenta_id_int = int(cuenta_id)
                    cuenta = Cuenta.objects.get(id=cuenta_id_int)
                    worksheet.write(f'A{row}', cuenta.nombre, normal_format)
                    worksheet.write(f'B{row}', float(saldo), money_format)
                    row += 1
                except (ValueError, Cuenta.DoesNotExist):
                    continue
        else:
            # Formato antiguo: {cuenta_id: saldo}
            for cuenta_id, saldo in corte.saldos_cuentas.items():
                try:
                    # Asegurar que cuenta_id sea un número
                    cuenta_id_int = int(cuenta_id)
                    cuenta = Cuenta.objects.get(id=cuenta_id_int)
                    worksheet.write(f'A{row}', cuenta.nombre, normal_format)
                    worksheet.write(f'B{row}', float(saldo), money_format)
                    row += 1
                except (ValueError, Cuenta.DoesNotExist):
                    continue
    
    # Gastos compartidos (si están disponibles)
    if isinstance(corte.saldos_cuentas, dict) and 'gastos_compartidos' in corte.saldos_cuentas:
        gastos_compartidos = corte.saldos_cuentas['gastos_compartidos']
        
        row += 2
        worksheet.merge_range(f'A{row}:D{row}', 'Gastos Compartidos del Mes', subtitle_format)
        row += 1
        
        worksheet.write(f'A{row}', 'Concepto', header_format)
        worksheet.write(f'B{row}', 'Monto', header_format)
        row += 1
        
        gc_resumen_data = [
            ['Total Gastos Compartidos', gastos_compartidos.get('total_gastos_compartidos_mes', 0)],
            ['Gastos Pagados por Usuario', gastos_compartidos.get('gastos_pagados_por_usuario', 0)],
            ['Pagos Realizados por Usuario', gastos_compartidos.get('pagos_usuario_mes', 0)],
            ['Total Saldo Pendiente', gastos_compartidos.get('total_saldo_pendiente', 0)],
        ]
        
        for concepto, monto in gc_resumen_data:
            worksheet.write(f'A{row}', concepto, normal_format)
            worksheet.write(f'B{row}', monto, money_format)
            row += 1
        
        # Detalle de saldos por grupo
        if 'saldos_pendientes_grupos' in gastos_compartidos:
            row += 2
            worksheet.merge_range(f'A{row}:E{row}', 'Saldos Pendientes por Grupo', subtitle_format)
            row += 1
            
            worksheet.write(f'A{row}', 'Grupo', header_format)
            worksheet.write(f'B{row}', 'Total Debido', header_format)
            worksheet.write(f'C{row}', 'Total Pagado', header_format)
            worksheet.write(f'D{row}', 'Saldo Pendiente', header_format)
            worksheet.write(f'E{row}', 'Miembros', header_format)
            row += 1
            
            for grupo_id, datos_grupo in gastos_compartidos['saldos_pendientes_grupos'].items():
                worksheet.write(f'A{row}', datos_grupo['nombre'], normal_format)
                worksheet.write(f'B{row}', datos_grupo['total_debido'], money_format)
                worksheet.write(f'C{row}', datos_grupo['total_pagado'], money_format)
                worksheet.write(f'D{row}', datos_grupo['saldo_pendiente'], money_format)
                worksheet.write(f'E{row}', datos_grupo['cantidad_miembros'], normal_format)
                row += 1
    
    # Información adicional
    row += 2
    worksheet.merge_range(f'A{row}:D{row}', 'Información Adicional', subtitle_format)
    row += 1
    
    worksheet.write(f'A{row}', 'Fecha de corte:', normal_format)
    worksheet.write(f'B{row}', corte.fecha_corte.strftime('%d/%m/%Y'), normal_format)
    row += 1
    
    worksheet.write(f'A{row}', 'Mantener saldos:', normal_format)
    worksheet.write(f'B{row}', 'Sí' if corte.mantener_saldos else 'No', normal_format)
    row += 1
    
    worksheet.write(f'A{row}', 'Fecha de creación:', normal_format)
    worksheet.write(f'B{row}', corte.fecha_creacion.strftime('%d/%m/%Y %H:%M'), normal_format)
    
    # Ajustar ancho de columnas
    worksheet.set_column('A:A', 25)
    worksheet.set_column('B:B', 15)
    worksheet.set_column('C:E', 15)
    
    workbook.close()
    output.seek(0)
    
    # Crear la respuesta HTTP
    response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return response
